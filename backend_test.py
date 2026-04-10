#!/usr/bin/env python3
"""
CodHER Hackathon Platform - Backend API Testing
Tests all backend APIs with admin credentials and Excel import flow
"""
import requests
import sys
import json
import time
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook

class CodHERAPITester:
    def __init__(self, base_url="https://eval-portal-6.preview.emergentagent.com"):
        self.base_url = base_url
        self.token = None
        self.tests_run = 0
        self.tests_passed = 0
        self.import_id = None

    def run_test(self, name, method, endpoint, expected_status, data=None, files=None):
        """Run a single API test"""
        url = f"{self.base_url}/{endpoint}"
        headers = {'Content-Type': 'application/json'}
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'
        
        # Remove Content-Type for file uploads
        if files:
            headers.pop('Content-Type', None)

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                if files:
                    response = requests.post(url, files=files, headers=headers)
                else:
                    response = requests.post(url, json=data, headers=headers)
            elif method == 'PUT':
                response = requests.put(url, json=data, headers=headers)
            elif method == 'DELETE':
                response = requests.delete(url, headers=headers)

            success = response.status_code == expected_status
            if success:
                self.tests_passed += 1
                print(f"✅ Passed - Status: {response.status_code}")
                try:
                    return success, response.json()
                except:
                    return success, {}
            else:
                print(f"❌ Failed - Expected {expected_status}, got {response.status_code}")
                try:
                    error_data = response.json()
                    print(f"   Error: {error_data}")
                except:
                    print(f"   Response: {response.text[:200]}")
                return False, {}

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            return False, {}

    def test_health_check(self):
        """Test health endpoint"""
        return self.run_test("Health Check", "GET", "api/health", 200)

    def test_admin_login(self):
        """Test admin login with correct credentials"""
        success, response = self.run_test(
            "Admin Login",
            "POST",
            "api/auth/login",
            200,
            data={"email": "admin@codher.com", "password": "Ganesh73005"}
        )
        if success and 'access_token' in response:
            self.token = response['access_token']
            print(f"   Admin user: {response.get('user', {}).get('username')} ({response.get('user', {}).get('role')})")
            return True
        return False

    def test_dashboard_stats(self):
        """Test dashboard stats API"""
        return self.run_test("Dashboard Stats", "GET", "api/dashboard/stats", 200)

    def create_test_excel_file(self):
        """Create a test Excel file with sample data"""
        wb = Workbook()
        
        # Teams sheet
        teams_ws = wb.active
        teams_ws.title = "teams"
        teams_headers = ["team_id", "team_name", "team_lead_name", "team_lead_email", "team_lead_mobile", 
                        "college_name", "project_title", "project_domain", "member_2_name", "member_2_email",
                        "member_3_name", "member_3_email", "status"]
        teams_ws.append(teams_headers)
        
        # Sample team data
        teams_data = [
            ["T001", "Code Warriors", "Alice Johnson", "alice@test.com", "9876543210", 
             "Tech University", "AI Chat Bot", "AI/ML", "Bob Smith", "bob@test.com",
             "Carol Davis", "carol@test.com", "Active"],
            ["T002", "Data Miners", "David Wilson", "david@test.com", "9876543211",
             "Data College", "Blockchain Voting", "Blockchain", "Eve Brown", "eve@test.com",
             "Frank Miller", "frank@test.com", "Active"]
        ]
        for row in teams_data:
            teams_ws.append(row)
        
        # Mentors sheet
        mentors_ws = wb.create_sheet("mentors")
        mentors_headers = ["mentor_id", "mentor_name", "mentor_email", "mentor_mobile", 
                          "expertise", "organization", "max_team_capacity", "status"]
        mentors_ws.append(mentors_headers)
        
        # Sample mentor data
        mentors_data = [
            ["M001", "Dr. Sarah Tech", "sarah@mentor.com", "9876543220", 
             "AI/ML, Python", "Tech Corp", "3", "Active"],
            ["M002", "Prof. John Code", "john@mentor.com", "9876543221",
             "Web Development, React", "Code Institute", "2", "Active"]
        ]
        for row in mentors_data:
            mentors_ws.append(row)
        
        # Deadlines sheet
        deadlines_ws = wb.create_sheet("deadlines")
        deadlines_headers = ["round_name", "submission_deadline", "evaluation_deadline", "grace_period_minutes"]
        deadlines_ws.append(deadlines_headers)
        
        deadlines_data = [
            ["Round 1", "2024-12-31 23:59:59", "2025-01-05 23:59:59", "30"],
            ["Round 2", "2025-01-15 23:59:59", "2025-01-20 23:59:59", "30"],
            ["Round 3", "2025-02-01 23:59:59", "2025-02-05 23:59:59", "30"]
        ]
        for row in deadlines_data:
            deadlines_ws.append(row)
        
        # Rubrics sheet
        rubrics_ws = wb.create_sheet("rubrics")
        rubrics_headers = ["round_name", "category_name", "max_marks", "description"]
        rubrics_ws.append(rubrics_headers)
        
        rubrics_data = [
            ["Round 1", "Innovation", "20", "Creativity and uniqueness"],
            ["Round 1", "Technical Implementation", "20", "Code quality and functionality"],
            ["Round 1", "Presentation", "20", "Communication and demo"],
            ["Round 1", "Problem Solving", "20", "Approach to solving the problem"],
            ["Round 1", "Team Collaboration", "20", "Teamwork and coordination"]
        ]
        for row in rubrics_data:
            rubrics_ws.append(row)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def test_excel_upload(self):
        """Test Excel file upload"""
        excel_data = self.create_test_excel_file()
        files = {'file': ('test_data.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, response = self.run_test(
            "Excel Upload",
            "POST", 
            "api/import/upload",
            200,
            files=files
        )
        
        if success and 'import_id' in response:
            self.import_id = response['import_id']
            preview = response.get('preview', {})
            print(f"   Teams: {preview.get('teams_count', 0)}, Mentors: {preview.get('mentors_count', 0)}")
            print(f"   Errors: {len(preview.get('errors', []))}, Warnings: {len(preview.get('warnings', []))}")
            return True
        return False

    def test_import_confirm(self):
        """Test import confirmation"""
        if not self.import_id:
            print("❌ No import_id available for confirmation")
            return False
        
        success, response = self.run_test(
            "Import Confirm",
            "POST",
            f"api/import/confirm/{self.import_id}",
            200,
            data={"send_emails": False}  # Don't send emails in testing
        )
        
        if success:
            created = response.get('created', {})
            print(f"   Created - Teams: {created.get('teams', 0)}, Mentors: {created.get('mentors', 0)}, Users: {created.get('users', 0)}")
            return True
        return False

    def test_teams_api(self):
        """Test teams management APIs"""
        success, response = self.run_test("Get Teams", "GET", "api/teams", 200)
        if success:
            teams = response.get('teams', [])
            print(f"   Found {len(teams)} teams")
            return len(teams) > 0
        return False

    def test_mentors_api(self):
        """Test mentors management APIs"""
        success, response = self.run_test("Get Mentors", "GET", "api/mentors", 200)
        if success:
            mentors = response.get('mentors', [])
            print(f"   Found {len(mentors)} mentors")
            return len(mentors) > 0
        return False

    def test_auto_assign_mentors(self):
        """Test auto-assign mentors functionality"""
        return self.run_test("Auto Assign Mentors", "POST", "api/mapping/auto-assign", 200)

    def test_evaluations_api(self):
        """Test evaluations APIs"""
        return self.run_test("Get Evaluations", "GET", "api/evaluations", 200)

    def test_release_status(self):
        """Test release status API"""
        return self.run_test("Get Release Status", "GET", "api/release-status", 200)

    def test_deadlines_api(self):
        """Test deadlines API"""
        return self.run_test("Get Deadlines", "GET", "api/deadlines", 200)

    def test_settings_api(self):
        """Test settings API"""
        return self.run_test("Get Settings", "GET", "api/settings", 200)

    def test_email_logs(self):
        """Test email logs API"""
        return self.run_test("Get Email Logs", "GET", "api/email/logs", 200)

    def test_chat_rooms(self):
        """Test chat rooms API"""
        return self.run_test("Get Chat Rooms", "GET", "api/chat/rooms", 200)

def main():
    print("🚀 Starting CodHER Platform Backend API Tests")
    print("=" * 60)
    
    tester = CodHERAPITester()
    
    # Core functionality tests
    tests = [
        ("Health Check", tester.test_health_check),
        ("Admin Login", tester.test_admin_login),
        ("Dashboard Stats", tester.test_dashboard_stats),
        ("Excel Upload", tester.test_excel_upload),
        ("Import Confirm", tester.test_import_confirm),
        ("Teams API", tester.test_teams_api),
        ("Mentors API", tester.test_mentors_api),
        ("Auto Assign", tester.test_auto_assign_mentors),
        ("Evaluations API", tester.test_evaluations_api),
        ("Release Status", tester.test_release_status),
        ("Deadlines API", tester.test_deadlines_api),
        ("Settings API", tester.test_settings_api),
        ("Email Logs", tester.test_email_logs),
        ("Chat Rooms", tester.test_chat_rooms),
    ]
    
    failed_tests = []
    
    for test_name, test_func in tests:
        try:
            success = test_func()
            if not success:
                failed_tests.append(test_name)
        except Exception as e:
            print(f"❌ {test_name} - Exception: {str(e)}")
            failed_tests.append(test_name)
    
    # Print results
    print("\n" + "=" * 60)
    print(f"📊 Test Results: {tester.tests_passed}/{tester.tests_run} passed")
    
    if failed_tests:
        print(f"❌ Failed tests: {', '.join(failed_tests)}")
        return 1
    else:
        print("✅ All tests passed!")
        return 0

if __name__ == "__main__":
    sys.exit(main())