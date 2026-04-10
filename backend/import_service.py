import re
import random
import string
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
import httpx

REQUIRED_TEAM_HEADERS = ["team_id", "team_name", "team_lead_name", "team_lead_email", "team_lead_mobile"]
REQUIRED_MENTOR_HEADERS = ["mentor_id", "mentor_name", "mentor_email", "max_team_capacity"]
REQUIRED_DEADLINE_HEADERS = ["round_name", "submission_deadline"]
REQUIRED_RUBRIC_HEADERS = ["round_name", "category_name", "max_marks"]

def convert_google_sheets_url(url: str) -> str:
    """Convert Google Sheets URL to export XLSX URL."""
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        raise ValueError("Invalid Google Sheets URL")
    sheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

async def download_google_sheet(url: str) -> bytes:
    """Download Google Sheet as XLSX bytes."""
    export_url = convert_google_sheets_url(url)
    async with httpx.AsyncClient(follow_redirects=True, timeout=30) as client:
        response = await client.get(export_url)
        response.raise_for_status()
        return response.content

def generate_password(identifier: str) -> str:
    """Generate readable password: identifier + random digits."""
    digits = ''.join(random.choices(string.digits, k=4))
    clean_id = re.sub(r'[^a-zA-Z0-9]', '', identifier)[:8]
    return f"{clean_id}{digits}"

def validate_email_format(email: str) -> bool:
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, str(email))) if email else False

def validate_mobile(mobile: str) -> bool:
    if not mobile:
        return True
    return bool(re.match(r'^[0-9+\-\s()]{7,15}$', str(mobile)))

def parse_workbook(file_bytes: bytes):
    """Parse multi-sheet Excel workbook and return structured data with validation."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    result = {
        "teams": [], "mentors": [], "deadlines": [], "settings": [],
        "errors": [], "warnings": []
    }
    
    # Parse teams
    if "teams" in wb.sheetnames:
        ws = wb["teams"]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        missing = [h for h in REQUIRED_TEAM_HEADERS if h not in headers]
        if missing:
            result["errors"].append(f"Teams sheet missing required columns: {', '.join(missing)}")
        else:
            seen_ids = set()
            seen_emails = set()
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = dict(zip(headers, [str(v).strip() if v is not None else "" for v in row]))
                if not row_dict.get("team_id"):
                    continue
                row_errors = []
                if row_dict["team_id"] in seen_ids:
                    row_errors.append(f"Row {row_num}: Duplicate team_id '{row_dict['team_id']}'")
                seen_ids.add(row_dict["team_id"])
                email = row_dict.get("team_lead_email", "")
                if not validate_email_format(email):
                    row_errors.append(f"Row {row_num}: Invalid email '{email}'")
                if email in seen_emails:
                    row_errors.append(f"Row {row_num}: Duplicate email '{email}'")
                seen_emails.add(email)
                if not validate_mobile(row_dict.get("team_lead_mobile")):
                    row_errors.append(f"Row {row_num}: Invalid mobile '{row_dict.get('team_lead_mobile')}'")
                if row_errors:
                    result["errors"].extend(row_errors)
                else:
                    result["teams"].append(row_dict)
    else:
        result["errors"].append("Missing required sheet: 'teams'")
    
    # Parse mentors
    if "mentors" in wb.sheetnames:
        ws = wb["mentors"]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        missing = [h for h in REQUIRED_MENTOR_HEADERS if h not in headers]
        if missing:
            result["errors"].append(f"Mentors sheet missing required columns: {', '.join(missing)}")
        else:
            seen_ids = set()
            seen_emails = set()
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = dict(zip(headers, [str(v).strip() if v is not None else "" for v in row]))
                if not row_dict.get("mentor_id"):
                    continue
                row_errors = []
                if row_dict["mentor_id"] in seen_ids:
                    row_errors.append(f"Row {row_num}: Duplicate mentor_id '{row_dict['mentor_id']}'")
                seen_ids.add(row_dict["mentor_id"])
                email = row_dict.get("mentor_email", "")
                if not validate_email_format(email):
                    row_errors.append(f"Row {row_num}: Invalid email '{email}'")
                if email in seen_emails:
                    row_errors.append(f"Row {row_num}: Duplicate email '{email}'")
                seen_emails.add(email)
                # Validate max_team_capacity - handle both int and float strings
                capacity_str = row_dict.get("max_team_capacity", "").strip()
                if capacity_str:
                    try:
                        # Try to convert to float first (handles "5.0"), then to int
                        capacity_val = int(float(capacity_str))
                        if capacity_val <= 0:
                            row_errors.append(f"Row {row_num}: max_team_capacity must be a positive number")
                        else:
                            row_dict["max_team_capacity"] = str(capacity_val)
                    except (ValueError, TypeError):
                        row_errors.append(f"Row {row_num}: max_team_capacity must be numeric (got '{capacity_str}')")
                else:
                    row_errors.append(f"Row {row_num}: max_team_capacity is required")
                if row_errors:
                    result["errors"].extend(row_errors)
                else:
                    result["mentors"].append(row_dict)
    else:
        result["warnings"].append("Missing optional sheet: 'mentors'")
    
    # Parse deadlines
    if "deadlines" in wb.sheetnames:
        ws = wb["deadlines"]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, [str(v).strip() if v is not None else "" for v in row]))
            if row_dict.get("round_name"):
                result["deadlines"].append(row_dict)
    
    # Rubrics are now hardcoded in frontend - skip parsing
    
    # Parse settings
    if "settings" in wb.sheetnames:
        ws = wb["settings"]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, [str(v).strip() if v is not None else "" for v in row]))
            if row_dict.get("key"):
                result["settings"].append(row_dict)
    
    wb.close()
    return result
