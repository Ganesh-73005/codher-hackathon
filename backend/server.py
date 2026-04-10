"""
CodHER Hackathon Management Platform - Main Server
FastAPI backend with RBAC, Excel import, evaluations, chat, email automation
"""
import os
import re
import json
import random
import string
import asyncio
from datetime import datetime, timedelta
from typing import Optional, List
from io import BytesIO
from contextlib import asynccontextmanager
from dateutil import parser as date_parser
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form, WebSocket, WebSocketDisconnect, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from bson import ObjectId
from pydantic import BaseModel, Field
from dotenv import load_dotenv

load_dotenv()

from database import db, init_db
from auth import (
    hash_password, verify_password, create_access_token, create_refresh_token,
    decode_token, get_current_user, require_role, security
)
from email_service import (
    send_credential_email, send_evaluation_status_email,
    send_results_release_email, send_custom_email, send_email
)
from import_service import parse_workbook, generate_password, download_google_sheet, convert_google_sheets_url
from websocket_manager import manager

# ─── Helpers ───────────────────────────────────────────────────────────
def serialize_doc(doc):
    """Serialize MongoDB document for JSON response."""
    if doc is None:
        return None
    if isinstance(doc, list):
        return [serialize_doc(d) for d in doc]
    result = {}
    for key, value in doc.items():
        if isinstance(value, ObjectId):
            result[key] = str(value)
        elif isinstance(value, datetime):
            result[key] = value.isoformat()
        elif isinstance(value, dict):
            result[key] = serialize_doc(value)
        elif isinstance(value, list):
            result[key] = [serialize_doc(v) if isinstance(v, dict) else str(v) if isinstance(v, ObjectId) else v.isoformat() if isinstance(v, datetime) else v for v in value]
        else:
            result[key] = value
    return result

# ─── App Setup ─────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    await ensure_admin()
    yield

app = FastAPI(title="CodHER Hackathon Platform", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

async def ensure_admin():
    """Create admin account if it doesn't exist."""
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@codher.com")
    admin_username = os.environ.get("ADMIN_USERNAME", "Ganesh")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Ganesh73005")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "username": admin_username,
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "role": "admin",
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
        })
        print(f"Admin account created: {admin_email}")

# ─── Pydantic Models ──────────────────────────────────────────────────
class LoginRequest(BaseModel):
    email: str
    password: str

class RefreshRequest(BaseModel):
    refresh_token: str

class EvaluationSubmit(BaseModel):
    team_id: str
    round_name: str
    scores: List[dict]  # [{category_name, score, comment}]
    feedback: str

class SubmissionUpdate(BaseModel):
    round_name: str
    ppt_link: Optional[str] = ""      # GDrive .pdf link (all rounds)
    github_link: Optional[str] = ""   # GitHub repo link (Round 2, 3)
    video_link: Optional[str] = ""    # GDrive video link (Round 3 only)
    team_id: Optional[str] = None     # Allow admin to specify team_id

class EmailCompose(BaseModel):
    recipients: List[str]  # list of emails
    subject: str
    body_html: str

class MentorAssignment(BaseModel):
    team_id: str
    mentor_email: str

# ─── Auth Routes ───────────────────────────────────────────────────────
@app.post("/api/auth/login")
async def login(req: LoginRequest):
    user = await db.users.find_one({"email": req.email})
    if not user or not verify_password(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    access_token = create_access_token({"sub": str(user["_id"]), "role": user["role"]})
    refresh_token = create_refresh_token({"sub": str(user["_id"]), "role": user["role"]})
    
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "user": serialize_doc({
            "_id": user["_id"],
            "username": user.get("username"),
            "email": user["email"],
            "role": user["role"],
            "team_id": user.get("team_id"),
            "mentor_id": user.get("mentor_id"),
        })
    }

@app.post("/api/auth/refresh")
async def refresh_token(req: RefreshRequest):
    payload = decode_token(req.refresh_token)
    if not payload or payload.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Invalid refresh token")
    access_token = create_access_token({"sub": payload["sub"], "role": payload["role"]})
    return {"access_token": access_token}

@app.get("/api/auth/me")
async def get_me(user=Depends(get_current_user)):
    return {"user": serialize_doc({
        "_id": user["_id"],
        "username": user.get("username"),
        "email": user["email"],
        "role": user["role"],
        "team_id": user.get("team_id"),
        "mentor_id": user.get("mentor_id"),
    })}

# ─── Import Routes (Admin) ────────────────────────────────────────────
@app.post("/api/import/upload")
async def import_upload(file: UploadFile = File(...), user=Depends(require_role("admin"))):
    """Upload Excel file and return preview data."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are supported")
    
    content = await file.read()
    result = parse_workbook(content)
    
    # Store temporarily for confirm step
    import_id = str(ObjectId())
    await db.import_sessions.insert_one({
        "_id": ObjectId(import_id),
        "data": result,
        "file_bytes": content,
        "status": "preview",
        "created_at": datetime.utcnow(),
        "created_by": str(user["_id"]),
    })
    
    return {
        "import_id": import_id,
        "preview": {
            "teams_count": len(result["teams"]),
            "mentors_count": len(result["mentors"]),
            "deadlines_count": len(result["deadlines"]),
            "settings_count": len(result["settings"]),
            "errors": result["errors"],
            "warnings": result["warnings"],
            "teams": result["teams"][:20],
            "mentors": result["mentors"][:20],
            "deadlines": result["deadlines"],
            "settings": result["settings"],
        }
    }

@app.post("/api/import/google-sheets")
async def import_google_sheets(url: str = Body(..., embed=True), user=Depends(require_role("admin"))):
    """Import from Google Sheets URL."""
    try:
        content = await download_google_sheet(url)
        result = parse_workbook(content)
        
        import_id = str(ObjectId())
        await db.import_sessions.insert_one({
            "_id": ObjectId(import_id),
            "data": result,
            "file_bytes": content,
            "source_url": url,
            "status": "preview",
            "created_at": datetime.utcnow(),
            "created_by": str(user["_id"]),
        })
        
        return {
            "import_id": import_id,
            "preview": {
                "teams_count": len(result["teams"]),
                "mentors_count": len(result["mentors"]),
                "deadlines_count": len(result["deadlines"]),
                "settings_count": len(result["settings"]),
                "errors": result["errors"],
                "warnings": result["warnings"],
                "teams": result["teams"][:20],
                "mentors": result["mentors"][:20],
                "deadlines": result["deadlines"],
                "settings": result["settings"],
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to import from Google Sheets: {str(e)}")

@app.post("/api/import/confirm/{import_id}")
async def import_confirm(import_id: str, send_emails: bool = Body(True, embed=True), user=Depends(require_role("admin"))):
    """Confirm import and create users/teams/mentors."""
    session = await db.import_sessions.find_one({"_id": ObjectId(import_id)})
    if not session:
        raise HTTPException(status_code=404, detail="Import session not found")
    
    data = session["data"]
    created = {"teams": 0, "mentors": 0, "users": 0, "deadlines": 0, "email_sent": 0, "email_failed": 0}
    errors = []
    
    # Import mentors first
    for mentor_data in data.get("mentors", []):
        try:
            existing = await db.mentors.find_one({"mentor_id": mentor_data["mentor_id"]})
            if existing:
                await db.mentors.update_one({"mentor_id": mentor_data["mentor_id"]}, {"$set": {
                    "mentor_name": mentor_data.get("mentor_name", ""),
                    "mentor_mobile": mentor_data.get("mentor_mobile", ""),
                    "expertise": mentor_data.get("expertise", ""),
                    "organization": mentor_data.get("organization", ""),
                    "max_team_capacity": int(float(mentor_data.get("max_team_capacity", 3))),
                    "status": mentor_data.get("status", "Active"),
                    "updated_at": datetime.utcnow(),
                }})
            else:
                await db.mentors.insert_one({
                    "mentor_id": mentor_data["mentor_id"],
                    "mentor_name": mentor_data.get("mentor_name", ""),
                    "mentor_email": mentor_data["mentor_email"],
                    "mentor_mobile": mentor_data.get("mentor_mobile", ""),
                    "expertise": mentor_data.get("expertise", ""),
                    "organization": mentor_data.get("organization", ""),
                    "max_team_capacity": int(float(mentor_data.get("max_team_capacity", 3))),
                    "status": mentor_data.get("status", "Active"),
                    "created_at": datetime.utcnow(),
                    "updated_at": datetime.utcnow(),
                })
                created["mentors"] += 1
            
            # Create user account for mentor
            existing_user = await db.users.find_one({"email": mentor_data["mentor_email"]})
            if not existing_user:
                password = generate_password(mentor_data["mentor_email"].split("@")[0])
                await db.users.insert_one({
                    "username": mentor_data["mentor_email"].split("@")[0],
                    "email": mentor_data["mentor_email"],
                    "password_hash": hash_password(password),
                    "plain_password_temp": password,  # Store for credential retrieval
                    "role": "mentor",
                    "mentor_id": mentor_data["mentor_id"],
                    "created_at": datetime.utcnow(),
                    "updated_at": datetime.utcnow(),
                })
                created["users"] += 1
                
                if send_emails:
                    result = await send_credential_email(
                        mentor_data["mentor_email"],
                        mentor_data["mentor_email"].split("@")[0],
                        password,
                        "mentor",
                        db
                    )
                    if result["status"] == "sent":
                        created["email_sent"] += 1
                    else:
                        created["email_failed"] += 1
        except Exception as e:
            errors.append(f"Mentor {mentor_data.get('mentor_id')}: {str(e)}")
    
    # Import teams
    for team_data in data.get("teams", []):
        try:
            existing = await db.teams.find_one({"team_id": team_data["team_id"]})
            members = []
            for i in range(2, 5):
                name = team_data.get(f"member_{i}_name", "")
                email = team_data.get(f"member_{i}_email", "")
                if name:
                    members.append({"name": name, "email": email})
            
            team_doc = {
                "team_id": team_data["team_id"],
                "team_name": team_data.get("team_name", ""),
                "college_name": team_data.get("college_name", ""),
                "project_title": team_data.get("project_title", ""),
                "project_domain": team_data.get("project_domain", ""),
                "team_lead_name": team_data.get("team_lead_name", ""),
                "team_lead_email": team_data.get("team_lead_email", ""),
                "team_lead_mobile": team_data.get("team_lead_mobile", ""),
                "members": members,
                "assigned_mentor_email": team_data.get("assigned_mentor_email", ""),
                "status": team_data.get("status", "Active"),
                "round_1_submission_link": team_data.get("round_1_submission_link", ""),
                "round_2_submission_link": team_data.get("round_2_submission_link", ""),
                "round_3_ppt_link": team_data.get("round_3_ppt_link", ""),
                "round_3_video_link": team_data.get("round_3_video_link", ""),
                "updated_at": datetime.utcnow(),
            }
            
            if existing:
                await db.teams.update_one({"team_id": team_data["team_id"]}, {"$set": team_doc})
            else:
                team_doc["created_at"] = datetime.utcnow()
                await db.teams.insert_one(team_doc)
                created["teams"] += 1
            
            # Create user account for team lead
            existing_user = await db.users.find_one({"email": team_data["team_lead_email"]})
            if not existing_user:
                password = generate_password(team_data["team_lead_email"].split("@")[0])
                await db.users.insert_one({
                    "username": team_data["team_lead_email"].split("@")[0],
                    "email": team_data["team_lead_email"],
                    "password_hash": hash_password(password),
                    "plain_password_temp": password,  # Store for credential retrieval
                    "role": "team",
                    "team_id": team_data["team_id"],
                    "created_at": datetime.utcnow(),
                    "updated_at": datetime.utcnow(),
                })
                created["users"] += 1
                
                if send_emails:
                    result = await send_credential_email(
                        team_data["team_lead_email"],
                        team_data["team_lead_email"].split("@")[0],
                        password,
                        "team",
                        db
                    )
                    if result["status"] == "sent":
                        created["email_sent"] += 1
                    else:
                        created["email_failed"] += 1
        except Exception as e:
            errors.append(f"Team {team_data.get('team_id')}: {str(e)}")
    
    # Import deadlines
    for dl in data.get("deadlines", []):
        try:
            await db.deadlines.update_one(
                {"round_name": dl["round_name"]},
                {"$set": {
                    "round_name": dl["round_name"],
                    "submission_deadline": dl.get("submission_deadline", ""),
                    "evaluation_deadline": dl.get("evaluation_deadline", ""),
                    "grace_period_minutes": int(float(dl.get("grace_period_minutes", 0) or 0)),
                    "updated_at": datetime.utcnow(),
                }},
                upsert=True
            )
            created["deadlines"] += 1
        except Exception as e:
            errors.append(f"Deadline {dl.get('round_name')}: {str(e)}")
    
    # Rubrics are now hardcoded in frontend - skip import
    
    # Import settings
    for setting in data.get("settings", []):
        try:
            await db.settings.update_one(
                {"key": setting["key"]},
                {"$set": {"key": setting["key"], "value": setting["value"], "updated_at": datetime.utcnow()}},
                upsert=True
            )
        except Exception:
            pass
    
    # Auto-create chat rooms for assigned team-mentor pairs
    all_teams = await db.teams.find({"assigned_mentor_email": {"$ne": "", "$exists": True}}).to_list(500)
    chat_rooms_created = 0
    for team in all_teams:
        if team.get("assigned_mentor_email"):
            existing_room = await db.chat_rooms.find_one({
                "team_id": team["team_id"],
                "mentor_email": team["assigned_mentor_email"],
            })
            if not existing_room:
                await db.chat_rooms.insert_one({
                    "team_id": team["team_id"],
                    "team_name": team.get("team_name", ""),
                    "mentor_email": team["assigned_mentor_email"],
                    "participants": [team["team_lead_email"], team["assigned_mentor_email"]],
                    "created_at": datetime.utcnow(),
                })
                chat_rooms_created += 1
    created["chat_rooms"] = chat_rooms_created
    
    # Update import session
    await db.import_sessions.update_one(
        {"_id": ObjectId(import_id)},
        {"$set": {"status": "confirmed", "confirmed_at": datetime.utcnow()}}
    )
    
    # Log sync
    await db.sync_logs.insert_one({
        "type": "import",
        "import_id": import_id,
        "created": created,
        "errors": errors,
        "timestamp": datetime.utcnow(),
        "triggered_by": str(user["_id"]),
    })
    
    return {"status": "success", "created": created, "errors": errors}

# ─── Teams Routes ──────────────────────────────────────────────────────
@app.get("/api/teams")
async def get_teams(
    search: str = "", status: str = "", page: int = 1, limit: int = 50,
    user=Depends(get_current_user)
):
    query = {}
    if user["role"] == "team":
        query["team_id"] = user.get("team_id")
    elif user["role"] == "mentor":
        mentor = await db.mentors.find_one({"mentor_id": user.get("mentor_id")})
        if mentor is not None:
            # Get traditionally assigned teams
            assigned = await db.teams.find({"assigned_mentor_email": mentor["mentor_email"]}).to_list(100)
            team_ids = [t["team_id"] for t in assigned]

            # Get round-wise assigned teams
            round_mappings = await db.round_mentor_assignments.find({"mentor_id": user.get("mentor_id")}).to_list(100)
            round_team_ids = [m["team_id"] for m in round_mappings]

            # Combine both lists (unique team IDs)
            all_team_ids = list(set(team_ids + round_team_ids))

            if all_team_ids:
                query["team_id"] = {"$in": all_team_ids}
            else:
                # No teams assigned - return empty
                query["team_id"] = {"$in": []}
    
    if search:
        query["$or"] = [
            {"team_name": {"$regex": search, "$options": "i"}},
            {"team_id": {"$regex": search, "$options": "i"}},
            {"team_lead_name": {"$regex": search, "$options": "i"}},
            {"project_title": {"$regex": search, "$options": "i"}},
        ]
    if status and status != "all":
        query["status"] = status
    
    total = await db.teams.count_documents(query)
    teams = await db.teams.find(query).skip((page - 1) * limit).limit(limit).to_list(limit)
    
    # Batch-fetch evaluation statuses and submissions for all teams (scalable for 200+)
    team_ids = [t["team_id"] for t in teams]
    
    # Batch fetch evaluations
    eval_pipeline = [
        {"$match": {"team_id": {"$in": team_ids}}},
        {"$group": {"_id": {"team_id": "$team_id", "round_name": "$round_name"}, "status": {"$first": "$status"}}}
    ]
    eval_results = {}
    async for doc in db.evaluations.aggregate(eval_pipeline):
        key = f"{doc['_id']['team_id']}_{doc['_id']['round_name']}"
        eval_results[key] = doc["status"]
    
    # Batch fetch submissions
    sub_pipeline = [
        {"$match": {"team_id": {"$in": team_ids}}},
        {"$group": {"_id": {"team_id": "$team_id", "round_name": "$round_name"}, "has_sub": {"$first": True}}}
    ]
    sub_results = set()
    async for doc in db.submissions.aggregate(sub_pipeline):
        sub_results.add(f"{doc['_id']['team_id']}_{doc['_id']['round_name']}")
    
    # Batch fetch round mappings if user is mentor
    round_mappings_by_team = {}
    if user["role"] == "mentor":
        mentor_id = user.get("mentor_id")
        mappings = await db.round_mentor_assignments.find({"mentor_id": mentor_id}).to_list(500)
        for mapping in mappings:
            team_id = mapping["team_id"]
            if team_id not in round_mappings_by_team:
                round_mappings_by_team[team_id] = []
            round_mappings_by_team[team_id].append(mapping["round_name"])

    enriched = []
    for team in teams:
        team_data = serialize_doc(team)
        for round_name in ["Round 1", "Round 2", "Round 3"]:
            key = f"{team['team_id']}_{round_name}"
            team_data[f"{round_name.lower().replace(' ', '_')}_eval_status"] = eval_results.get(key, "pending")
            team_data[f"{round_name.lower().replace(' ', '_')}_submitted"] = key in sub_results
        # Include team mobile number for mentors
        team_data["team_lead_mobile"] = team.get("team_lead_mobile", "")
        # Include round assignments for this team
        team_data["assigned_rounds"] = round_mappings_by_team.get(team["team_id"], [])
        enriched.append(team_data)

    return {"teams": enriched, "total": total, "page": page, "limit": limit}

@app.get("/api/teams/{team_id}")
async def get_team(team_id: str, user=Depends(get_current_user)):
    team = await db.teams.find_one({"team_id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    if user["role"] == "team" and user.get("team_id") != team_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return {"team": serialize_doc(team)}

# ─── Mentors Routes ───────────────────────────────────────────────────
@app.get("/api/mentors")
async def get_mentors(search: str = "", user=Depends(get_current_user)):
    query = {}
    if user["role"] == "mentor":
        query["mentor_id"] = user.get("mentor_id")
    if search:
        query["$or"] = [
            {"mentor_name": {"$regex": search, "$options": "i"}},
            {"mentor_email": {"$regex": search, "$options": "i"}},
            {"expertise": {"$regex": search, "$options": "i"}},
        ]
    
    mentors = await db.mentors.find(query).to_list(100)
    
    # Batch-fetch assignment counts and eval completion (scalable for 20+ mentors)
    mentor_emails = [m["mentor_email"] for m in mentors]
    
    # Batch count assigned teams per mentor
    assign_pipeline = [
        {"$match": {"assigned_mentor_email": {"$in": mentor_emails}}},
        {"$group": {"_id": "$assigned_mentor_email", "count": {"$sum": 1}, "team_ids": {"$push": "$team_id"}}}
    ]
    assign_counts = {}
    mentor_team_map = {}
    async for doc in db.teams.aggregate(assign_pipeline):
        assign_counts[doc["_id"]] = doc["count"]
        mentor_team_map[doc["_id"]] = doc["team_ids"]
    
    # Batch count evaluations per mentor
    all_team_ids = []
    for tids in mentor_team_map.values():
        all_team_ids.extend(tids)
    
    eval_done_pipeline = [
        {"$match": {"team_id": {"$in": all_team_ids}, "status": "evaluated"}},
        {"$group": {"_id": "$team_id", "count": {"$sum": 1}}}
    ]
    eval_done_map = {}
    async for doc in db.evaluations.aggregate(eval_done_pipeline):
        eval_done_map[doc["_id"]] = doc["count"]
    
    enriched = []
    for mentor in mentors:
        m = serialize_doc(mentor)
        email = mentor["mentor_email"]
        assigned = assign_counts.get(email, 0)
        m["assigned_teams_count"] = assigned
        team_ids_for_mentor = mentor_team_map.get(email, [])
        total_evals = len(team_ids_for_mentor) * 3
        completed_evals = sum(eval_done_map.get(tid, 0) for tid in team_ids_for_mentor)
        m["eval_completion"] = f"{completed_evals}/{total_evals}" if total_evals > 0 else "0/0"
        enriched.append(m)
    
    return {"mentors": enriched}

@app.get("/api/mentors/{mentor_id}")
async def get_mentor(mentor_id: str, user=Depends(get_current_user)):
    mentor = await db.mentors.find_one({"mentor_id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    return {"mentor": serialize_doc(mentor)}

@app.patch("/api/mentors/{mentor_id}/capacity")
async def update_mentor_capacity(
    mentor_id: str,
    capacity: int = Body(..., embed=True),
    user=Depends(require_role("admin"))
):
    """Update mentor's maximum team capacity."""
    if capacity < 0 or capacity > 50:
        raise HTTPException(status_code=400, detail="Capacity must be between 0 and 50")

    mentor = await db.mentors.find_one({"mentor_id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")

    await db.mentors.update_one(
        {"mentor_id": mentor_id},
        {"$set": {"max_team_capacity": capacity, "updated_at": datetime.utcnow()}}
    )

    return {"status": "success", "message": f"Capacity updated to {capacity}", "new_capacity": capacity}

class MentorCreate(BaseModel):
    mentor_name: str
    mentor_email: str
    mentor_mobile: str
    expertise: str = ""
    organization: str = ""
    max_team_capacity: int = 3
    send_email: bool = True

@app.post("/api/mentors")
async def create_mentor(mentor: MentorCreate, user=Depends(require_role("admin"))):
    """Manually create a new mentor."""
    # Check if mentor already exists
    existing_mentor = await db.mentors.find_one({"mentor_email": mentor.mentor_email})
    if existing_mentor:
        raise HTTPException(status_code=400, detail="Mentor with this email already exists")

    # Generate mentor_id
    count = await db.mentors.count_documents({})
    mentor_id = f"M{str(count + 1).zfill(3)}"

    # Check if ID already exists (in case of concurrent creation)
    while await db.mentors.find_one({"mentor_id": mentor_id}):
        count += 1
        mentor_id = f"M{str(count + 1).zfill(3)}"

    # Create mentor document
    mentor_data = {
        "mentor_id": mentor_id,
        "mentor_name": mentor.mentor_name,
        "mentor_email": mentor.mentor_email,
        "mentor_mobile": mentor.mentor_mobile,
        "expertise": mentor.expertise,
        "organization": mentor.organization,
        "max_team_capacity": mentor.max_team_capacity,
        "status": "Active",
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }

    await db.mentors.insert_one(mentor_data)

    # Create user account
    password = generate_password(mentor.mentor_email.split("@")[0])
    await db.users.insert_one({
        "username": mentor.mentor_email.split("@")[0],
        "email": mentor.mentor_email,
        "password_hash": hash_password(password),
        "plain_password_temp": password,
        "role": "mentor",
        "mentor_id": mentor_id,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow(),
    })

    # Send credentials email if requested
    if mentor.send_email:
        try:
            await send_credential_email(
                mentor.mentor_email,
                mentor.mentor_email.split("@")[0],
                password,
                "mentor",
                db
            )
        except Exception as e:
            # Don't fail mentor creation if email fails
            print(f"Failed to send email: {e}")

    return {
        "status": "success",
        "message": f"Mentor {mentor_id} created successfully",
        "mentor_id": mentor_id,
        "password": password if not mentor.send_email else None
    }

@app.post("/api/teams/send-credentials")
async def send_team_credentials(team_id: str = Body(..., embed=True), user=Depends(require_role("admin"))):
    """Send credentials email to a specific team."""
    team = await db.teams.find_one({"team_id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    team_user = await db.users.find_one({"email": team["team_lead_email"], "role": "team"})
    if not team_user:
        raise HTTPException(status_code=404, detail="Team user account not found")

    password = team_user.get("plain_password_temp")
    if not password:
        raise HTTPException(status_code=400, detail="Password not available. User may have been created before password storage was implemented.")

    result = await send_credential_email(
        team["team_lead_email"],
        team_user.get("username", team["team_lead_email"].split("@")[0]),
        password,
        "team",
        db
    )

    if result["status"] == "sent":
        return {"status": "success", "message": f"Credentials sent to {team['team_lead_email']}"}
    else:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {result.get('error', 'Unknown error')}")

@app.post("/api/teams/send-all-credentials")
async def send_all_team_credentials(user=Depends(require_role("admin"))):
    """Send credentials email to all teams."""
    teams = await db.teams.find().to_list(1000)
    sent = 0
    failed = 0
    errors = []

    for team in teams:
        try:
            team_user = await db.users.find_one({"email": team["team_lead_email"], "role": "team"})
            if not team_user:
                continue

            password = team_user.get("plain_password_temp")
            if not password:
                errors.append(f"{team['team_id']}: Password not available")
                continue

            result = await send_credential_email(
                team["team_lead_email"],
                team_user.get("username", team["team_lead_email"].split("@")[0]),
                password,
                "team",
                db
            )

            if result["status"] == "sent":
                sent += 1
            else:
                failed += 1
                errors.append(f"{team['team_id']}: {result.get('error', 'Unknown')}")
        except Exception as e:
            failed += 1
            errors.append(f"{team['team_id']}: {str(e)}")

    return {
        "status": "success",
        "sent": sent,
        "failed": failed,
        "errors": errors[:10]  # Return first 10 errors
    }

@app.post("/api/mentors/send-credentials")
async def send_mentor_credentials(mentor_id: str = Body(..., embed=True), user=Depends(require_role("admin"))):
    """Send credentials email to a specific mentor."""
    mentor = await db.mentors.find_one({"mentor_id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")

    mentor_user = await db.users.find_one({"email": mentor["mentor_email"], "role": "mentor"})
    if not mentor_user:
        raise HTTPException(status_code=404, detail="Mentor user account not found")

    password = mentor_user.get("plain_password_temp")
    if not password:
        raise HTTPException(status_code=400, detail="Password not available. User may have been created before password storage was implemented.")

    result = await send_credential_email(
        mentor["mentor_email"],
        mentor_user.get("username", mentor["mentor_email"].split("@")[0]),
        password,
        "mentor",
        db
    )

    if result["status"] == "sent":
        return {"status": "success", "message": f"Credentials sent to {mentor['mentor_email']}"}
    else:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {result.get('error', 'Unknown error')}")

@app.post("/api/mentors/send-all-credentials")
async def send_all_mentor_credentials(user=Depends(require_role("admin"))):
    """Send credentials email to all mentors."""
    mentors = await db.mentors.find().to_list(100)
    sent = 0
    failed = 0
    errors = []

    for mentor in mentors:
        try:
            mentor_user = await db.users.find_one({"email": mentor["mentor_email"], "role": "mentor"})
            if not mentor_user:
                continue

            password = mentor_user.get("plain_password_temp")
            if not password:
                errors.append(f"{mentor['mentor_id']}: Password not available")
                continue

            result = await send_credential_email(
                mentor["mentor_email"],
                mentor_user.get("username", mentor["mentor_email"].split("@")[0]),
                password,
                "mentor",
                db
            )

            if result["status"] == "sent":
                sent += 1
            else:
                failed += 1
                errors.append(f"{mentor['mentor_id']}: {result.get('error', 'Unknown')}")
        except Exception as e:
            failed += 1
            errors.append(f"{mentor['mentor_id']}: {str(e)}")

    return {
        "status": "success",
        "sent": sent,
        "failed": failed,
        "errors": errors[:10]  # Return first 10 errors
    }

# ─── Round-wise Mentor Mapping Routes (Admin) ────────────────────────

@app.get("/api/round-mappings")
async def get_round_mappings(user=Depends(require_role("admin"))):
    """Get all round-wise mentor-team mappings."""
    mappings = await db.round_mentor_assignments.find().to_list(1000)
    return {"mappings": [serialize_doc(m) for m in mappings]}

@app.get("/api/round-mappings/mentor/{mentor_id}")
async def get_mentor_round_mappings(mentor_id: str, user=Depends(get_current_user)):
    """Get round mappings for a specific mentor."""
    mappings = await db.round_mentor_assignments.find({"mentor_id": mentor_id}).to_list(100)
    return {"mappings": [serialize_doc(m) for m in mappings]}

@app.get("/api/round-mappings/team/{team_id}")
async def get_team_round_mappings(team_id: str, user=Depends(get_current_user)):
    """Get round mappings for a specific team."""
    mappings = await db.round_mentor_assignments.find({"team_id": team_id}).to_list(10)
    return {"mappings": [serialize_doc(m) for m in mappings]}

class RoundMappingCreate(BaseModel):
    mentor_id: str
    team_ids: list[str]
    rounds: list[str]  # e.g., ["Round 1", "Round 2"]

@app.post("/api/round-mappings")
async def create_round_mappings(mapping: RoundMappingCreate, user=Depends(require_role("admin"))):
    """Create round-wise mentor-team mappings."""
    mentor = await db.mentors.find_one({"mentor_id": mapping.mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")

    created_count = 0
    for team_id in mapping.team_ids:
        team = await db.teams.find_one({"team_id": team_id})
        if not team:
            continue

        for round_name in mapping.rounds:
            # Check if mapping already exists
            existing = await db.round_mentor_assignments.find_one({
                "mentor_id": mapping.mentor_id,
                "team_id": team_id,
                "round_name": round_name
            })

            if not existing:
                await db.round_mentor_assignments.insert_one({
                    "mentor_id": mapping.mentor_id,
                    "mentor_email": mentor["mentor_email"],
                    "mentor_name": mentor.get("mentor_name", ""),
                    "team_id": team_id,
                    "team_name": team.get("team_name", ""),
                    "round_name": round_name,
                    "status": "active",
                    "created_at": datetime.utcnow(),
                    "updated_at": datetime.utcnow()
                })
                created_count += 1

    return {"status": "success", "created": created_count, "message": f"Created {created_count} round mappings"}

@app.delete("/api/round-mappings/{mapping_id}")
async def delete_round_mapping(mapping_id: str, user=Depends(require_role("admin"))):
    """Delete a specific round mapping."""
    from bson import ObjectId
    try:
        result = await db.round_mentor_assignments.delete_one({"_id": ObjectId(mapping_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Mapping not found")
        return {"status": "success", "message": "Mapping deleted"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/api/round-mappings/bulk")
async def delete_bulk_round_mappings(
    mentor_id: str = Query(None),
    team_id: str = Query(None),
    round_name: str = Query(None),
    user=Depends(require_role("admin"))
):
    """Delete multiple round mappings based on filters."""
    query = {}
    if mentor_id:
        query["mentor_id"] = mentor_id
    if team_id:
        query["team_id"] = team_id
    if round_name:
        query["round_name"] = round_name

    if not query:
        raise HTTPException(status_code=400, detail="At least one filter parameter required")

    result = await db.round_mentor_assignments.delete_many(query)
    return {"status": "success", "deleted": result.deleted_count}

# ─── Mapping Routes (Admin) ───────────────────────────────────────────
@app.post("/api/mapping/assign")
async def assign_mentor(assignment: MentorAssignment, user=Depends(require_role("admin"))):
    """Manually assign mentor to team."""
    team = await db.teams.find_one({"team_id": assignment.team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    mentor = await db.mentors.find_one({"mentor_email": assignment.mentor_email})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    # Check capacity
    assigned_count = await db.teams.count_documents({"assigned_mentor_email": assignment.mentor_email})
    if assigned_count >= mentor.get("max_team_capacity", 3):
        raise HTTPException(status_code=400, detail="Mentor at maximum capacity")
    
    old_mentor = team.get("assigned_mentor_email", "")
    await db.teams.update_one(
        {"team_id": assignment.team_id},
        {"$set": {"assigned_mentor_email": assignment.mentor_email, "updated_at": datetime.utcnow()}}
    )
    
    # Audit trail
    await db.assignment_logs.insert_one({
        "team_id": assignment.team_id,
        "old_mentor_email": old_mentor,
        "new_mentor_email": assignment.mentor_email,
        "assigned_by": str(user["_id"]),
        "timestamp": datetime.utcnow(),
    })
    
    # Auto-create chat room
    existing_room = await db.chat_rooms.find_one({
        "team_id": assignment.team_id,
        "mentor_email": assignment.mentor_email,
    })
    if not existing_room:
        await db.chat_rooms.insert_one({
            "team_id": assignment.team_id,
            "team_name": team.get("team_name", ""),
            "mentor_email": assignment.mentor_email,
            "participants": [team["team_lead_email"], assignment.mentor_email],
            "created_at": datetime.utcnow(),
        })
    
    return {"status": "success", "message": f"Team {assignment.team_id} assigned to {assignment.mentor_email}"}

@app.post("/api/mapping/auto-assign")
async def auto_assign_mentors(user=Depends(require_role("admin"))):
    """Auto-assign unassigned teams to mentors based on capacity."""
    unassigned = await db.teams.find({"$or": [{"assigned_mentor_email": ""}, {"assigned_mentor_email": None}]}).to_list(500)
    mentors = await db.mentors.find({"status": "Active"}).to_list(100)
    
    if not mentors:
        raise HTTPException(status_code=400, detail="No active mentors available")
    
    assignments = []
    for team in unassigned:
        # Find mentor with least assignments under capacity
        best_mentor = None
        min_count = float('inf')
        for mentor in mentors:
            count = await db.teams.count_documents({"assigned_mentor_email": mentor["mentor_email"]})
            capacity = int(float(mentor.get("max_team_capacity", 3)))
            if count < capacity and count < min_count:
                best_mentor = mentor
                min_count = count
        
        if best_mentor:
            await db.teams.update_one(
                {"team_id": team["team_id"]},
                {"$set": {"assigned_mentor_email": best_mentor["mentor_email"], "updated_at": datetime.utcnow()}}
            )
            await db.assignment_logs.insert_one({
                "team_id": team["team_id"],
                "old_mentor_email": "",
                "new_mentor_email": best_mentor["mentor_email"],
                "assigned_by": "auto",
                "timestamp": datetime.utcnow(),
            })
            # Auto-create chat room
            existing_room = await db.chat_rooms.find_one({
                "team_id": team["team_id"],
                "mentor_email": best_mentor["mentor_email"],
            })
            if not existing_room:
                await db.chat_rooms.insert_one({
                    "team_id": team["team_id"],
                    "team_name": team.get("team_name", ""),
                    "mentor_email": best_mentor["mentor_email"],
                    "participants": [team["team_lead_email"], best_mentor["mentor_email"]],
                    "created_at": datetime.utcnow(),
                })
            assignments.append({"team_id": team["team_id"], "mentor_email": best_mentor["mentor_email"]})
    
    return {"status": "success", "assignments": assignments, "count": len(assignments)}

@app.post("/api/mapping/bulk-reassign")
async def bulk_reassign(assignments: List[MentorAssignment], user=Depends(require_role("admin"))):
    """Bulk reassign mentors."""
    results = []
    for a in assignments:
        try:
            team = await db.teams.find_one({"team_id": a.team_id})
            if team is not None:
                old_mentor = team.get("assigned_mentor_email", "")
                await db.teams.update_one(
                    {"team_id": a.team_id},
                    {"$set": {"assigned_mentor_email": a.mentor_email, "updated_at": datetime.utcnow()}}
                )
                await db.assignment_logs.insert_one({
                    "team_id": a.team_id,
                    "old_mentor_email": old_mentor,
                    "new_mentor_email": a.mentor_email,
                    "assigned_by": str(user["_id"]),
                    "timestamp": datetime.utcnow(),
                })
                results.append({"team_id": a.team_id, "status": "success"})
        except Exception as e:
            results.append({"team_id": a.team_id, "status": "failed", "error": str(e)})
    return {"results": results}

# ─── Submissions Routes ───────────────────────────────────────────────
@app.get("/api/submissions")
async def get_submissions(team_id: str = "", round_name: str = "", user=Depends(get_current_user)):
    query = {}
    if user["role"] == "team":
        query["team_id"] = user.get("team_id")
    elif team_id:
        query["team_id"] = team_id
    if round_name:
        query["round_name"] = round_name
    
    submissions = await db.submissions.find(query).to_list(500)
    return {"submissions": serialize_doc(submissions)}

@app.post("/api/submissions")
async def create_submission(sub: SubmissionUpdate, user=Depends(get_current_user)):
    team_id = user.get("team_id")
    if user["role"] == "admin":
        team_id = sub.team_id or team_id
    if not team_id:
        raise HTTPException(status_code=400, detail="No team associated")
    
    # Check deadline (only for teams, not admin)
    if user["role"] == "team":
        deadline = await db.deadlines.find_one({"round_name": sub.round_name})
        if deadline and deadline.get("submission_deadline"):
            deadline_str = deadline["submission_deadline"]
            try:
                # Try parsing various date formats
                deadline_dt = date_parser.parse(deadline_str)
                current_time = datetime.utcnow()
                
                # Add grace period if available
                grace_minutes = int(float(deadline.get("grace_period_minutes", 0)))
                deadline_with_grace = deadline_dt + timedelta(minutes=grace_minutes)
                
                if current_time > deadline_with_grace:
                    grace_msg = f" (including {grace_minutes} min grace period)" if grace_minutes > 0 else ""
                    raise HTTPException(
                        status_code=403, 
                        detail=f"Submission deadline for {sub.round_name} has passed on {deadline_str}{grace_msg}"
                    )
            except ValueError:
                # If deadline format is invalid, allow submission but log warning
                print(f"Warning: Invalid deadline format for {sub.round_name}: {deadline_str}")
    
    # Validate round-specific fields
    errors = []
    round_config = {
        "Round 1": {"required": ["ppt_link"], "optional": []},
        "Round 2": {"required": ["ppt_link", "github_link"], "optional": []},
        "Round 3": {"required": ["ppt_link", "github_link", "video_link"], "optional": []},
    }
    config = round_config.get(sub.round_name, round_config["Round 1"])
    
    # Validate GDrive links
    if sub.ppt_link:
        if not re.match(r'https?://(drive\.google\.com|docs\.google\.com)', sub.ppt_link):
            errors.append("PPT link must be a Google Drive link")
    if sub.video_link:
        if not re.match(r'https?://(drive\.google\.com|docs\.google\.com|youtu)', sub.video_link):
            errors.append("Video link must be a Google Drive link")
    if sub.github_link:
        if not re.match(r'https?://(github\.com|gitlab\.com)', sub.github_link):
            errors.append("GitHub link must be a valid GitHub/GitLab URL")
    
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    
    submission_data = {
        "team_id": team_id,
        "round_name": sub.round_name,
        "ppt_link": sub.ppt_link or "",
        "github_link": sub.github_link or "",
        "video_link": sub.video_link or "",
        "updated_at": datetime.utcnow(),
    }
    
    existing = await db.submissions.find_one({"team_id": team_id, "round_name": sub.round_name})
    if existing:
        await db.submissions.update_one(
            {"_id": existing["_id"]},
            {"$set": submission_data}
        )
    else:
        submission_data["submitted_at"] = datetime.utcnow()
        await db.submissions.insert_one(submission_data)
    
    return {"status": "success", "message": f"Submission for {sub.round_name} saved"}

# ─── Evaluations Routes ───────────────────────────────────────────────
@app.get("/api/evaluations")
async def get_evaluations(team_id: str = "", round_name: str = "", mentor_id: str = "", user=Depends(get_current_user)):
    query = {}
    
    if user["role"] == "team":
        query["team_id"] = user.get("team_id")
    elif user["role"] == "mentor":
        query["mentor_id"] = user.get("mentor_id")
    
    if team_id:
        query["team_id"] = team_id
    if round_name:
        query["round_name"] = round_name
    if mentor_id and user["role"] != "mentor":
        query["mentor_id"] = mentor_id
    
    evals = await db.evaluations.find(query).to_list(500)
    result = []
    
    for ev in evals:
        ev_data = serialize_doc(ev)
        scores = await db.evaluation_scores.find({"evaluation_id": ev["_id"]}).to_list(10)
        
        if user["role"] == "team":
            # Check release flag
            round_key = f"release_{ev['round_name'].lower().replace(' ', '_')}"
            release_flag = await db.settings.find_one({"key": round_key})
            global_flag = await db.settings.find_one({"key": "release_global"})
            is_released = (
                (release_flag and release_flag.get("value") == "true") or
                (global_flag and global_flag.get("value") == "true")
            )
            
            if not is_released:
                # Redact scores for teams
                ev_data.pop("total_score", None)
                ev_data.pop("feedback", None)
                ev_data["scores"] = []
                ev_data["released"] = False
            else:
                ev_data["scores"] = serialize_doc(scores)
                ev_data["released"] = True
        else:
            ev_data["scores"] = serialize_doc(scores)
            ev_data["released"] = True

        # Get team name
        team = await db.teams.find_one({"team_id": ev["team_id"]})
        if team is not None:
            ev_data["team_name"] = team.get("team_name", "")
        
        result.append(ev_data)
    
    return {"evaluations": result}

@app.post("/api/evaluations")
async def submit_evaluation(ev: EvaluationSubmit, user=Depends(require_role("admin", "mentor"))):
    """Submit or update evaluation."""
    mentor_id = user.get("mentor_id")
    if user["role"] == "admin":
        mentor_id = "admin"

    if not mentor_id:
        raise HTTPException(status_code=400, detail="No mentor profile associated")

    # Validate team exists
    team = await db.teams.find_one({"team_id": ev.team_id})

    # Check if mentor is authorized to evaluate this team for this round
    if user["role"] == "mentor":
        # Check round-wise assignment
        round_mapping = await db.round_mentor_assignments.find_one({
            "mentor_id": mentor_id,
            "team_id": ev.team_id,
            "round_name": ev.round_name
        })

        # Also check traditional assignment (for backward compatibility)
        mentor = await db.mentors.find_one({"mentor_id": mentor_id})
        is_traditionally_assigned = False
        if mentor and team:
            is_traditionally_assigned = team.get("assigned_mentor_email") == mentor.get("mentor_email")

        if not round_mapping and not is_traditionally_assigned:
            raise HTTPException(
                status_code=403,
                detail=f"You are not assigned to evaluate {ev.team_id} for {ev.round_name}"
            )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Validate scores - hardcoded max scores per category
    CATEGORY_MAX_SCORES = {
        "Innovation & Creativity": 20,
        "Technical Implementation": 25,
        "Functionality & Completion": 20,
        "Presentation & Communication": 15,
        "Impact & Relevance to Track": 10,
        "Code Quality & Scalability": 10
    }

    total_score = 0
    for score in ev.scores:
        s = int(float(score.get("score", 0)))
        category_name = score.get("category_name", "")
        max_score = CATEGORY_MAX_SCORES.get(category_name, 20)  # Default to 20 if category not found

        if s < 0 or s > max_score:
            raise HTTPException(status_code=400, detail=f"Score for {category_name} must be 0-{max_score}")
        total_score += s
    
    # Check for existing evaluation
    existing = await db.evaluations.find_one({
        "team_id": ev.team_id,
        "round_name": ev.round_name,
        "mentor_id": mentor_id,
    })
    
    if existing:
        # Update
        await db.evaluations.update_one(
            {"_id": existing["_id"]},
            {"$set": {
                "total_score": total_score,
                "feedback": ev.feedback,
                "status": "evaluated",
                "updated_at": datetime.utcnow(),
            }}
        )
        eval_id = existing["_id"]
        # Remove old scores
        await db.evaluation_scores.delete_many({"evaluation_id": existing["_id"]})
    else:
        # Create new
        result = await db.evaluations.insert_one({
            "team_id": ev.team_id,
            "mentor_id": mentor_id,
            "round_name": ev.round_name,
            "total_score": total_score,
            "feedback": ev.feedback,
            "status": "evaluated",
            "submitted_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
        })
        eval_id = result.inserted_id
    
    # Insert scores
    for score in ev.scores:
        await db.evaluation_scores.insert_one({
            "evaluation_id": eval_id,
            "category_name": score["category_name"],
            "score": int(float(score.get("score", 0))),
            "max_score": 20,
            "comment": score.get("comment", ""),
        })
    
    # Send status-only email to team
    try:
        await send_evaluation_status_email(
            team["team_lead_email"],
            team["team_name"],
            ev.round_name,
            db
        )
    except Exception as e:
        print(f"Failed to send evaluation email: {e}")
    
    return {"status": "success", "total_score": total_score, "message": "Evaluation submitted"}

@app.get("/api/evaluations/export/excel")
async def export_evaluations_excel(round_name: str = "", user=Depends(require_role("admin"))):
    """Export evaluations to Excel."""
    query = {}
    if round_name:
        query["round_name"] = round_name

    evals = await db.evaluations.find(query).to_list(500)

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    # Excel sheet titles have max 31 chars and can't have special chars
    sheet_title = round_name[:31] if round_name else "All Evaluations"
    ws.title = sheet_title

    # Headers - Updated for new rubrics categories
    headers = [
        "Team ID", "Team Name", "Mentor ID", "Round",
        "Innovation (20)", "Technical (25)", "Functionality (20)",
        "Presentation (15)", "Impact (10)", "Code Quality (10)",
        "Total Score (100)", "Status", "Comments", "Submitted At"
    ]
    ws.append(headers)

    # Data
    for ev in evals:
        team = await db.teams.find_one({"team_id": ev["team_id"]})
        scores_list = await db.evaluation_scores.find({"evaluation_id": ev["_id"]}).to_list(10)

        scores_dict = {}
        for s in scores_list:
            scores_dict[s["category_name"]] = s.get("score", 0)

        # Format submitted_at properly
        submitted_str = ""
        if ev.get("submitted_at"):
            try:
                if isinstance(ev["submitted_at"], str):
                    submitted_str = ev["submitted_at"]
                else:
                    submitted_str = ev["submitted_at"].strftime("%Y-%m-%d %H:%M")
            except:
                submitted_str = str(ev.get("submitted_at", ""))

        ws.append([
            str(ev.get("team_id", "")),
            str(team.get("team_name", "")) if team else "",
            str(ev.get("mentor_id", "")),
            str(ev.get("round_name", "")),
            int(scores_dict.get("Innovation & Creativity", 0)),
            int(scores_dict.get("Technical Implementation", 0)),
            int(scores_dict.get("Functionality & Completion", 0)),
            int(scores_dict.get("Presentation & Communication", 0)),
            int(scores_dict.get("Impact & Relevance to Track", 0)),
            int(scores_dict.get("Code Quality & Scalability", 0)),
            int(ev.get("total_score", 0)),
            str(ev.get("status", "")),
            str(ev.get("feedback", "")),
            submitted_str
        ])

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"evaluations_{round_name.replace(' ', '_').lower()}.xlsx" if round_name else "evaluations_all.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/evaluations/export/pdf")
async def export_evaluations_pdf(round_name: str = "", user=Depends(require_role("admin"))):
    """Export evaluations to PDF."""
    query = {}
    if round_name:
        query["round_name"] = round_name

    evals = await db.evaluations.find(query).to_list(500)

    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    # Title
    title_text = f"CodHER Hackathon - Evaluations ({round_name})" if round_name else "CodHER Hackathon - All Evaluations"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    # Scores Summary Table - Updated for new rubrics
    data = [["Team", "Mentor", "Inn\n(20)", "Tech\n(25)", "Func\n(20)", "Pres\n(15)", "Imp\n(10)", "Code\n(10)", "Total"]]

    evals_with_teams = []
    for ev in evals:
        team = await db.teams.find_one({"team_id": ev["team_id"]})
        scores_list = await db.evaluation_scores.find({"evaluation_id": ev["_id"]}).to_list(10)

        scores_dict = {}
        for s in scores_list:
            scores_dict[s["category_name"]] = s.get("score", 0)

        team_name = str(team.get("team_name", ""))[:12] if team else ""
        mentor_id = str(ev.get("mentor_id", ""))[:6]

        data.append([
            team_name,
            mentor_id,
            str(int(scores_dict.get("Innovation & Creativity", 0))),
            str(int(scores_dict.get("Technical Implementation", 0))),
            str(int(scores_dict.get("Functionality & Completion", 0))),
            str(int(scores_dict.get("Presentation & Communication", 0))),
            str(int(scores_dict.get("Impact & Relevance to Track", 0))),
            str(int(scores_dict.get("Code Quality & Scalability", 0))),
            str(int(ev.get("total_score", 0)))
        ])

        # Store for comments section
        evals_with_teams.append({
            "team_name": team.get("team_name", "") if team else "",
            "mentor_id": ev.get("mentor_id", ""),
            "comments": ev.get("feedback", "")
        })

    # Create scores table with updated column widths
    table = Table(data, colWidths=[1.2*inch, 0.6*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a5568')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))

    # Comments Section
    if evals_with_teams:
        comment_style = ParagraphStyle(
            'CommentStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            leftIndent=10
        )

        heading_style = ParagraphStyle(
            'CommentHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER
        )

        comments_heading = Paragraph("Evaluation Comments", heading_style)
        elements.append(comments_heading)

        for item in evals_with_teams:
            if item["comments"]:
                team_header = Paragraph(
                    f"<b>{item['team_name']}</b> (Mentor: {item['mentor_id']})",
                    styles['Heading3']
                )
                elements.append(team_header)

                # Sanitize comments for PDF (escape special characters)
                clean_comments = str(item["comments"]).replace('<', '&lt;').replace('>', '&gt;').replace('&', '&amp;')
                comment_text = Paragraph(clean_comments, comment_style)
                elements.append(comment_text)
                elements.append(Spacer(1, 0.15*inch))

    doc.build(elements)

    output.seek(0)
    filename = f"evaluations_{round_name.replace(' ', '_').lower()}.pdf" if round_name else "evaluations_all.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ─── Results Release Routes (Admin) ───────────────────────────────────
@app.get("/api/release-status")
async def get_release_status(user=Depends(get_current_user)):
    """Get release flags for all rounds."""
    flags = {}
    for round_name in ["Round 1", "Round 2", "Round 3"]:
        key = f"release_{round_name.lower().replace(' ', '_')}"
        setting = await db.settings.find_one({"key": key})
        flags[round_name] = setting.get("value", "false") == "true" if setting else False
    
    global_setting = await db.settings.find_one({"key": "release_global"})
    flags["Global"] = global_setting.get("value", "false") == "true" if global_setting else False
    
    # Get release logs
    logs = await db.result_release_logs.find().sort("released_at", -1).to_list(50)
    
    return {"flags": flags, "logs": serialize_doc(logs)}

@app.post("/api/release-results")
async def release_results(
    round_name: str = Body(...),
    release: bool = Body(True),
    user=Depends(require_role("admin"))
):
    """Release or unreleased results for a round."""
    if round_name == "Global":
        key = "release_global"
    else:
        key = f"release_{round_name.lower().replace(' ', '_')}"
    
    await db.settings.update_one(
        {"key": key},
        {"$set": {"key": key, "value": "true" if release else "false", "updated_at": datetime.utcnow()}},
        upsert=True
    )
    
    # Log
    await db.result_release_logs.insert_one({
        "round_name": round_name,
        "action": "released" if release else "revoked",
        "released_by": str(user["_id"]),
        "released_by_name": user.get("username", "admin"),
        "released_at": datetime.utcnow(),
    })
    
    # Send emails to all teams if releasing
    if release:
        teams = await db.teams.find({"status": "Active"}).to_list(500)
        for team in teams:
            try:
                await send_results_release_email(
                    team["team_lead_email"],
                    team["team_name"],
                    round_name,
                    db
                )
            except Exception as e:
                print(f"Failed to send release email to {team['team_lead_email']}: {e}")
    
    return {"status": "success", "message": f"Results {'released' if release else 'revoked'} for {round_name}"}

# ─── Leaderboard Route ────────────────────────────────────────────────
@app.get("/api/leaderboard")
async def get_leaderboard(user=Depends(get_current_user)):
    """Get global leaderboard - only visible when global release is enabled."""
    # Check if global release is enabled
    global_setting = await db.settings.find_one({"key": "release_global"})
    is_global_released = global_setting.get("value", "false") == "true" if global_setting else False
    
    # Allow admin to view always, others only when released
    if user["role"] != "admin" and not is_global_released:
        return {"leaderboard": [], "released": False, "message": "Leaderboard will be visible when results are released"}
    
    # Get all teams with evaluations
    teams = await db.teams.find().to_list(500)
    leaderboard = []
    
    for team in teams:
        # Get all evaluations for this team
        evaluations = await db.evaluations.find({"team_id": team["team_id"]}).to_list(100)
        
        total_score = 0
        round_scores = {}
        rounds_evaluated = 0
        
        for ev in evaluations:
            round_name = ev["round_name"]
            round_total = ev.get("total_score", 0)
            
            # Store highest score for each round (in case of multiple evaluations)
            if round_name not in round_scores or round_total > round_scores[round_name]:
                round_scores[round_name] = round_total
        
        # Sum up all round scores
        total_score = sum(round_scores.values())
        rounds_evaluated = len(round_scores)
        
        if rounds_evaluated > 0:  # Only include teams with at least one evaluation
            leaderboard.append({
                "team_id": team["team_id"],
                "team_name": team["team_name"],
                "team_lead_name": team.get("team_lead_name", ""),
                "college": team.get("college", ""),
                "total_score": total_score,
                "rounds_evaluated": rounds_evaluated,
                "round_scores": round_scores,
                "assigned_mentor_email": team.get("assigned_mentor_email", ""),
            })
    
    # Sort by total score (descending)
    leaderboard.sort(key=lambda x: x["total_score"], reverse=True)
    
    # Add rank
    for idx, entry in enumerate(leaderboard):
        entry["rank"] = idx + 1
    
    return {
        "leaderboard": leaderboard,
        "released": is_global_released,
        "total_teams": len(leaderboard)
    }

# ─── Export Routes ────────────────────────────────────────────────────
@app.get("/api/leaderboard/export/excel")
async def export_leaderboard_excel(user=Depends(get_current_user)):
    """Export leaderboard to Excel."""
    # Get leaderboard data
    global_setting = await db.settings.find_one({"key": "release_global"})
    is_global_released = global_setting.get("value", "false") == "true" if global_setting else False

    if user["role"] != "admin" and not is_global_released:
        raise HTTPException(status_code=403, detail="Leaderboard not released yet")

    teams = await db.teams.find().to_list(500)
    leaderboard = []

    for team in teams:
        evaluations = await db.evaluations.find({"team_id": team["team_id"]}).to_list(100)
        total_score = 0
        round_scores = {}

        for ev in evaluations:
            round_name = ev["round_name"]
            round_total = ev.get("total_score", 0)
            if round_name not in round_scores or round_total > round_scores[round_name]:
                round_scores[round_name] = round_total

        total_score = sum(round_scores.values())

        if len(round_scores) > 0:
            leaderboard.append({
                "team_id": team["team_id"],
                "team_name": team["team_name"],
                "team_lead_name": team.get("team_lead_name", ""),
                "college": team.get("college", ""),
                "total_score": total_score,
                "round_1": round_scores.get("Round 1", 0),
                "round_2": round_scores.get("Round 2", 0),
                "round_3": round_scores.get("Round 3", 0),
            })

    leaderboard.sort(key=lambda x: x["total_score"], reverse=True)

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Leaderboard"

    # Headers
    headers = ["Rank", "Team ID", "Team Name", "Team Lead", "College", "Round 1", "Round 2", "Round 3", "Total Score"]
    ws.append(headers)

    # Data
    for idx, entry in enumerate(leaderboard):
        ws.append([
            idx + 1,
            str(entry.get("team_id", "")),
            str(entry.get("team_name", "")),
            str(entry.get("team_lead_name", "")),
            str(entry.get("college", "")),
            int(entry.get("round_1", 0)),
            int(entry.get("round_2", 0)),
            int(entry.get("round_3", 0)),
            int(entry.get("total_score", 0))
        ])

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leaderboard.xlsx"}
    )

@app.get("/api/leaderboard/export/pdf")
async def export_leaderboard_pdf(user=Depends(get_current_user)):
    """Export leaderboard to PDF."""
    # Get leaderboard data (same as Excel)
    global_setting = await db.settings.find_one({"key": "release_global"})
    is_global_released = global_setting.get("value", "false") == "true" if global_setting else False

    if user["role"] != "admin" and not is_global_released:
        raise HTTPException(status_code=403, detail="Leaderboard not released yet")

    teams = await db.teams.find().to_list(500)
    leaderboard = []

    for team in teams:
        evaluations = await db.evaluations.find({"team_id": team["team_id"]}).to_list(100)
        total_score = 0
        round_scores = {}

        for ev in evaluations:
            round_name = ev["round_name"]
            round_total = ev.get("total_score", 0)
            if round_name not in round_scores or round_total > round_scores[round_name]:
                round_scores[round_name] = round_total

        total_score = sum(round_scores.values())

        if len(round_scores) > 0:
            leaderboard.append({
                "team_name": team["team_name"],
                "team_lead_name": team.get("team_lead_name", ""),
                "college": team.get("college", ""),
                "total_score": total_score,
                "round_1": round_scores.get("Round 1", 0),
                "round_2": round_scores.get("Round 2", 0),
                "round_3": round_scores.get("Round 3", 0),
            })

    leaderboard.sort(key=lambda x: x["total_score"], reverse=True)

    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    # Title
    title = Paragraph("CodHER Hackathon - Global Leaderboard", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    # Table data
    data = [["Rank", "Team Name", "Lead", "R1", "R2", "R3", "Total"]]
    for idx, entry in enumerate(leaderboard[:50]):  # Limit to top 50 for PDF
        data.append([
            str(idx + 1),
            str(entry.get("team_name", ""))[:25],  # Truncate long names
            str(entry.get("team_lead_name", ""))[:20],
            str(entry.get("round_1", 0)),
            str(entry.get("round_2", 0)),
            str(entry.get("round_3", 0)),
            str(entry.get("total_score", 0))
        ])

    # Create table
    table = Table(data, colWidths=[0.6*inch, 2.2*inch, 1.5*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a5568')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=leaderboard.pdf"}
    )

# ─── Email Routes (Admin) ─────────────────────────────────────────────
@app.post("/api/email/send")
async def send_bulk_email(compose: EmailCompose, user=Depends(require_role("admin"))):
    """Send custom email to selected recipients."""
    results = {"sent": 0, "failed": 0, "errors": []}
    for email_addr in compose.recipients:
        try:
            result = await send_custom_email(email_addr, compose.subject, compose.body_html, db)
            if result["status"] == "sent":
                results["sent"] += 1
            else:
                results["failed"] += 1
                results["errors"].append(f"{email_addr}: {result.get('error')}")
        except Exception as e:
            results["failed"] += 1
            results["errors"].append(f"{email_addr}: {str(e)}")
    
    return results

@app.get("/api/email/logs")
async def get_email_logs(page: int = 1, limit: int = 50, user=Depends(require_role("admin"))):
    total = await db.email_logs.count_documents({})
    logs = await db.email_logs.find().sort("sent_at", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    return {"logs": serialize_doc(logs), "total": total}

# ─── Deadlines Routes ─────────────────────────────────────────────────
@app.get("/api/deadlines")
async def get_deadlines(user=Depends(get_current_user)):
    deadlines = await db.deadlines.find().to_list(10)
    return {"deadlines": serialize_doc(deadlines)}

@app.post("/api/deadlines")
async def update_deadline(
    round_name: str = Body(...),
    submission_deadline: str = Body(...),
    evaluation_deadline: str = Body(""),
    grace_period_minutes: int = Body(0),
    user=Depends(require_role("admin"))
):
    await db.deadlines.update_one(
        {"round_name": round_name},
        {"$set": {
            "round_name": round_name,
            "submission_deadline": submission_deadline,
            "evaluation_deadline": evaluation_deadline,
            "grace_period_minutes": grace_period_minutes,
            "updated_at": datetime.utcnow(),
        }},
        upsert=True
    )
    return {"status": "success"}

# ─── Rubrics Routes ───────────────────────────────────────────────────
@app.get("/api/rubrics")
async def get_rubrics(round_name: str = "", user=Depends(get_current_user)):
    query = {}
    if round_name:
        query["round_name"] = round_name
    rubrics = await db.rubrics.find(query).to_list(50)
    return {"rubrics": serialize_doc(rubrics)}

# ─── Dashboard Stats ──────────────────────────────────────────────────
@app.get("/api/dashboard/stats")
async def get_dashboard_stats(user=Depends(get_current_user)):
    """Get dashboard statistics based on role."""
    stats = {}
    
    if user["role"] == "admin":
        stats["total_teams"] = await db.teams.count_documents({})
        stats["active_teams"] = await db.teams.count_documents({"status": "Active"})
        stats["total_mentors"] = await db.mentors.count_documents({})
        stats["active_mentors"] = await db.mentors.count_documents({"status": "Active"})
        stats["total_evaluations"] = await db.evaluations.count_documents({})
        stats["completed_evaluations"] = await db.evaluations.count_documents({"status": "evaluated"})
        stats["pending_evaluations"] = (stats["active_teams"] * 3) - stats["completed_evaluations"]
        stats["total_submissions"] = await db.submissions.count_documents({})
        stats["emails_sent"] = await db.email_logs.count_documents({"status": "sent"})
        stats["emails_failed"] = await db.email_logs.count_documents({"status": "failed"})
        
        # Unassigned teams
        stats["unassigned_teams"] = await db.teams.count_documents({
            "$or": [{"assigned_mentor_email": ""}, {"assigned_mentor_email": None}, {"assigned_mentor_email": {"$exists": False}}]
        })
        
    elif user["role"] == "mentor":
        mentor = await db.mentors.find_one({"mentor_id": user.get("mentor_id")})
        if mentor is not None:
            assigned = await db.teams.find({"assigned_mentor_email": mentor["mentor_email"]}).to_list(100)
            stats["assigned_teams"] = len(assigned)
            stats["max_capacity"] = mentor.get("max_team_capacity", 3)
            team_ids = [t["team_id"] for t in assigned]
            # Batch query evaluations
            completed = await db.evaluations.count_documents({
                "team_id": {"$in": team_ids},
                "mentor_id": user.get("mentor_id"),
                "status": "evaluated"
            })
            total_possible = len(team_ids) * 3
            stats["completed_evaluations"] = completed
            stats["pending_evaluations"] = total_possible - completed

            # Get round-wise team assignments
            round_mappings = await db.round_mentor_assignments.find({"mentor_id": user.get("mentor_id")}).to_list(100)
            stats["round_assignments"] = {}
            for mapping in round_mappings:
                round_name = mapping["round_name"]
                if round_name not in stats["round_assignments"]:
                    stats["round_assignments"][round_name] = []
                stats["round_assignments"][round_name].append({
                    "team_id": mapping["team_id"],
                    "team_name": mapping.get("team_name", "")
                })
    
    elif user["role"] == "team":
        team_id = user.get("team_id")
        team = await db.teams.find_one({"team_id": team_id})
        if team is not None:
            stats["team_name"] = team.get("team_name", "")
            stats["mentor_email"] = team.get("assigned_mentor_email", "")
            # Get mentor details
            if stats["mentor_email"]:
                mentor = await db.mentors.find_one({"mentor_email": stats["mentor_email"]})
                if mentor is not None:
                    stats["mentor_name"] = mentor.get("mentor_name", "")
                    stats["mentor_expertise"] = mentor.get("expertise", "")
                    stats["mentor_mobile"] = mentor.get("mentor_mobile", "")

            # Get round-wise mentor assignments
            round_mappings = await db.round_mentor_assignments.find({"team_id": team_id}).to_list(10)
            stats["round_mentors"] = {}
            for mapping in round_mappings:
                round_name = mapping["round_name"]
                if round_name not in stats["round_mentors"]:
                    stats["round_mentors"][round_name] = []
                stats["round_mentors"][round_name].append({
                    "mentor_id": mapping["mentor_id"],
                    "mentor_name": mapping.get("mentor_name", ""),
                    "mentor_email": mapping["mentor_email"]
                })
            
            # Submission status per round
            stats["rounds"] = {}
            for r in ["Round 1", "Round 2", "Round 3"]:
                sub = await db.submissions.find_one({"team_id": team_id, "round_name": r})
                ev = await db.evaluations.find_one({"team_id": team_id, "round_name": r})
                
                round_key = f"release_{r.lower().replace(' ', '_')}"
                release_flag = await db.settings.find_one({"key": round_key})
                global_flag = await db.settings.find_one({"key": "release_global"})
                is_released = (
                    (release_flag and release_flag.get("value") == "true") or
                    (global_flag and global_flag.get("value") == "true")
                )
                
                if is_released and ev:
                    status_text = "Final Results Released"
                elif ev and ev.get("status") == "evaluated":
                    status_text = "Evaluated"
                elif ev:
                    status_text = "Under Review"
                elif sub:
                    status_text = "Submitted"
                else:
                    status_text = "Not Submitted"
                
                stats["rounds"][r] = {
                    "submission_status": "Submitted" if sub else "Not Submitted",
                    "eval_status": status_text,
                    "is_released": is_released,
                }
        
        # Get deadlines
        deadlines = await db.deadlines.find().to_list(10)
        stats["deadlines"] = serialize_doc(deadlines)
    
    return {"stats": stats}

# ─── Chat Routes ───────────────────────────────────────────────────────
@app.get("/api/chat/rooms")
async def get_chat_rooms(user=Depends(get_current_user)):
    """Get chat rooms for user."""
    if user["role"] == "admin":
        # Admin can see all rooms (real + virtual)
        existing_rooms = await db.chat_rooms.find().to_list(100)
        existing_pairs = {(room.get("team_id"), room.get("mentor_email")) for room in existing_rooms}

        # Get all teams
        teams = await db.teams.find().to_list(500)

        # Start with existing rooms
        rooms = list(existing_rooms)

        # Add virtual rooms for team-mentor pairs without existing rooms
        for team in teams:
            # Traditional assignment
            if team.get("assigned_mentor_email"):
                pair = (team["team_id"], team["assigned_mentor_email"])
                if pair not in existing_pairs:
                    mentor = await db.mentors.find_one({"mentor_email": team["assigned_mentor_email"]})
                    if mentor:
                        rooms.append({
                            "_id": f"virtual_{team['team_id']}_{team['assigned_mentor_email']}",
                            "team_id": team["team_id"],
                            "team_name": team.get("team_name", ""),
                            "mentor_email": team["assigned_mentor_email"],
                            "mentor_name": mentor.get("mentor_name", ""),
                            "participants": [team["team_lead_email"], team["assigned_mentor_email"]],
                            "is_virtual": True,
                            "created_at": None
                        })
                        existing_pairs.add(pair)

            # Round-wise assignments
            round_mappings = await db.round_mentor_assignments.find({"team_id": team["team_id"]}).to_list(20)
            for mapping in round_mappings:
                pair = (team["team_id"], mapping["mentor_email"])
                if pair not in existing_pairs:
                    mentor = await db.mentors.find_one({"mentor_email": mapping["mentor_email"]})
                    if mentor:
                        rooms.append({
                            "_id": f"virtual_{team['team_id']}_{mapping['mentor_email']}",
                            "team_id": team["team_id"],
                            "team_name": team.get("team_name", ""),
                            "mentor_email": mapping["mentor_email"],
                            "mentor_name": mentor.get("mentor_name", ""),
                            "participants": [team["team_lead_email"], mapping["mentor_email"]],
                            "is_virtual": True,
                            "created_at": None
                        })
                        existing_pairs.add(pair)
    elif user["role"] == "mentor":
        mentor = await db.mentors.find_one({"mentor_id": user.get("mentor_id")})
        if mentor is not None:
            # Get existing chat rooms
            existing_rooms = await db.chat_rooms.find({"participants": mentor["mentor_email"]}).to_list(100)
            existing_team_ids = {room["team_id"] for room in existing_rooms if "team_id" in room}

            # Get ALL teams assigned to this mentor
            traditional_teams = await db.teams.find({"assigned_mentor_email": mentor["mentor_email"]}).to_list(100)
            round_mappings = await db.round_mentor_assignments.find({"mentor_id": user.get("mentor_id")}).to_list(100)
            round_team_ids = [m["team_id"] for m in round_mappings]
            round_teams = await db.teams.find({"team_id": {"$in": round_team_ids}}).to_list(100) if round_team_ids else []

            all_teams = {t["team_id"]: t for t in traditional_teams + round_teams}.values()

            # Start with existing rooms
            rooms = list(existing_rooms)

            # Add virtual rooms for teams without existing rooms
            for team in all_teams:
                if team["team_id"] not in existing_team_ids:
                    rooms.append({
                        "_id": f"virtual_{team['team_id']}_{mentor['mentor_email']}",
                        "team_id": team["team_id"],
                        "team_name": team.get("team_name", ""),
                        "mentor_email": mentor["mentor_email"],
                        "participants": [team["team_lead_email"], mentor["mentor_email"]],
                        "is_virtual": True,
                        "created_at": None
                    })
        else:
            rooms = []
    elif user["role"] == "team":
        team = await db.teams.find_one({"team_id": user.get("team_id")})
        if team is not None:
            # Get existing chat rooms
            existing_rooms = await db.chat_rooms.find({"participants": team["team_lead_email"]}).to_list(100)
            existing_mentor_emails = {room.get("mentor_email") for room in existing_rooms if room.get("mentor_email")}

            # Get ALL mentors assigned to this team
            mentor_emails = set()
            if team.get("assigned_mentor_email"):
                mentor_emails.add(team["assigned_mentor_email"])

            # Get round-wise mentors
            round_mappings = await db.round_mentor_assignments.find({"team_id": team["team_id"]}).to_list(20)
            for mapping in round_mappings:
                mentor_emails.add(mapping["mentor_email"])

            # Start with existing rooms
            rooms = list(existing_rooms)

            # Add virtual rooms for mentors without existing rooms
            for mentor_email in mentor_emails:
                if mentor_email not in existing_mentor_emails:
                    mentor = await db.mentors.find_one({"mentor_email": mentor_email})
                    if mentor:
                        rooms.append({
                            "_id": f"virtual_{team['team_id']}_{mentor_email}",
                            "team_id": team["team_id"],
                            "team_name": team.get("team_name", ""),
                            "mentor_email": mentor_email,
                            "mentor_name": mentor.get("mentor_name", ""),
                            "participants": [team["team_lead_email"], mentor_email],
                            "is_virtual": True,
                            "created_at": None
                        })
        else:
            rooms = []
    else:
        rooms = []
    
    # Enrich with last message
    enriched = []
    for room in rooms:
        r = serialize_doc(room) if not room.get("is_virtual") else room

        # Skip message enrichment for virtual rooms
        if room.get("is_virtual"):
            r["last_message"] = None
            r["unread_count"] = 0
        else:
            last_msg = await db.chat_messages.find({"room_id": str(room["_id"])}).sort("timestamp", -1).limit(1).to_list(1)
            if last_msg:
                r["last_message"] = serialize_doc(last_msg[0])
            unread = await db.chat_messages.count_documents({
                "room_id": str(room["_id"]),
                "read_by": {"$nin": [user["email"]]}
            })
            r["unread_count"] = unread
        enriched.append(r)

    return {"rooms": enriched}

@app.post("/api/chat/rooms")
async def create_or_get_chat_room(
    team_id: str = Body(...),
    mentor_email: str = Body(...),
    user=Depends(get_current_user)
):
    """Create or get existing chat room between team and mentor."""
    team = await db.teams.find_one({"team_id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    existing = await db.chat_rooms.find_one({
        "team_id": team_id,
        "mentor_email": mentor_email,
    })

    if existing:
        return {"room": serialize_doc(existing)}

    # Get all round-wise mentors assigned to this team
    round_mappings = await db.round_mentor_assignments.find({"team_id": team_id}).to_list(20)
    mentor_emails = set([mentor_email])  # Primary mentor
    for mapping in round_mappings:
        mentor_emails.add(mapping["mentor_email"])

    # Build participants list: team lead + all mentors
    participants = [team["team_lead_email"]] + list(mentor_emails)

    room = {
        "team_id": team_id,
        "team_name": team.get("team_name", ""),
        "mentor_email": mentor_email,
        "participants": participants,
        "created_at": datetime.utcnow(),
    }
    result = await db.chat_rooms.insert_one(room)
    room["_id"] = result.inserted_id
    return {"room": serialize_doc(room)}

@app.get("/api/chat/messages/{room_id}")
async def get_chat_messages(room_id: str, page: int = 1, limit: int = 50, user=Depends(get_current_user)):
    """Get chat messages for a room."""
    messages = await db.chat_messages.find({"room_id": room_id}).sort("timestamp", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    messages.reverse()  # Oldest first
    
    # Mark as read
    await db.chat_messages.update_many(
        {"room_id": room_id, "read_by": {"$nin": [user["email"]]}},
        {"$addToSet": {"read_by": user["email"]}}
    )
    
    return {"messages": serialize_doc(messages)}

@app.post("/api/chat/messages")
async def send_chat_message(
    room_id: str = Body(...),
    content: str = Body(...),
    user=Depends(get_current_user)
):
    """Send a chat message (REST fallback for WebSocket)."""
    message = {
        "room_id": room_id,
        "sender_email": user["email"],
        "sender_name": user.get("username", user["email"]),
        "sender_role": user["role"],
        "content": content,
        "timestamp": datetime.utcnow(),
        "read_by": [user["email"]],
    }
    result = await db.chat_messages.insert_one(message)
    message["_id"] = result.inserted_id
    
    # Broadcast via WebSocket
    await manager.broadcast(room_id, serialize_doc(message))
    
    return {"message": serialize_doc(message)}

# ─── WebSocket Chat ────────────────────────────────────────────────────
@app.websocket("/api/ws/chat/{room_id}")
async def websocket_chat(websocket: WebSocket, room_id: str, token: str = Query(None)):
    """WebSocket endpoint for real-time chat."""
    # Authenticate
    if not token:
        await websocket.close(code=4001)
        return
    
    payload = decode_token(token)
    if not payload:
        await websocket.close(code=4001)
        return
    
    user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
    if not user:
        await websocket.close(code=4001)
        return
    
    user_info = {
        "email": user["email"],
        "username": user.get("username", ""),
        "role": user["role"],
    }
    
    await manager.connect(websocket, room_id, user_info)
    
    try:
        while True:
            data = await websocket.receive_json()
            
            if data.get("type") == "message":
                message = {
                    "room_id": room_id,
                    "sender_email": user["email"],
                    "sender_name": user.get("username", user["email"]),
                    "sender_role": user["role"],
                    "content": data.get("content", ""),
                    "timestamp": datetime.utcnow(),
                    "read_by": [user["email"]],
                }
                result = await db.chat_messages.insert_one(message)
                message["_id"] = result.inserted_id
                
                await manager.broadcast(room_id, {
                    **serialize_doc(message),
                    "type": "message",
                })
            
            elif data.get("type") == "typing":
                await manager.broadcast(room_id, {
                    "type": "typing",
                    "sender_email": user["email"],
                    "sender_name": user.get("username", ""),
                })
    
    except WebSocketDisconnect:
        manager.disconnect(websocket, room_id)
    except Exception:
        manager.disconnect(websocket, room_id)

# ─── Sync Routes (Admin) ──────────────────────────────────────────────
@app.post("/api/sync/trigger")
async def trigger_sync(url: str = Body(..., embed=True), user=Depends(require_role("admin"))):
    """Manually trigger sync from Google Sheets."""
    try:
        content = await download_google_sheet(url)
        result = parse_workbook(content)
        
        # Only update dynamic fields
        updated = {"teams_updated": 0, "deadlines_updated": 0}
        
        for team_data in result.get("teams", []):
            existing = await db.teams.find_one({"team_id": team_data["team_id"]})
            if existing:
                update_fields = {}
                for field in ["round_1_submission_link", "round_2_submission_link", "round_3_ppt_link", "round_3_video_link", "status"]:
                    if team_data.get(field) and team_data[field] != existing.get(field):
                        update_fields[field] = team_data[field]
                if update_fields:
                    update_fields["updated_at"] = datetime.utcnow()
                    await db.teams.update_one({"team_id": team_data["team_id"]}, {"$set": update_fields})
                    updated["teams_updated"] += 1
        
        for dl in result.get("deadlines", []):
            await db.deadlines.update_one(
                {"round_name": dl["round_name"]},
                {"$set": {
                    "submission_deadline": dl.get("submission_deadline", ""),
                    "evaluation_deadline": dl.get("evaluation_deadline", ""),
                    "grace_period_minutes": int(float(dl.get("grace_period_minutes", 0) or 0)),
                    "updated_at": datetime.utcnow(),
                }},
                upsert=True
            )
            updated["deadlines_updated"] += 1
        
        # Log sync
        await db.sync_logs.insert_one({
            "type": "manual_sync",
            "source_url": url,
            "updated": updated,
            "timestamp": datetime.utcnow(),
            "triggered_by": str(user["_id"]),
            "status": "success",
        })
        
        return {"status": "success", "updated": updated}
    except Exception as e:
        await db.sync_logs.insert_one({
            "type": "manual_sync",
            "source_url": url,
            "error": str(e),
            "timestamp": datetime.utcnow(),
            "triggered_by": str(user["_id"]),
            "status": "failed",
        })
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/sync/logs")
async def get_sync_logs(user=Depends(require_role("admin"))):
    logs = await db.sync_logs.find().sort("timestamp", -1).to_list(50)
    return {"logs": serialize_doc(logs)}

# ─── Settings Routes ──────────────────────────────────────────────────
@app.get("/api/settings")
async def get_settings(user=Depends(get_current_user)):
    settings = await db.settings.find().to_list(50)
    return {"settings": serialize_doc(settings)}

@app.post("/api/settings")
async def update_settings(key: str = Body(...), value: str = Body(...), user=Depends(require_role("admin"))):
    await db.settings.update_one(
        {"key": key},
        {"$set": {"key": key, "value": value, "updated_at": datetime.utcnow()}},
        upsert=True
    )
    return {"status": "success"}

# ─── Data Management (Admin) ──────────────────────────────────────────
@app.post("/api/admin/truncate")
async def truncate_data(confirm: str = Body(..., embed=True), user=Depends(require_role("admin"))):
    """Truncate all data except admin account. Requires confirm='DELETE_ALL_DATA'."""
    if confirm != "DELETE_ALL_DATA":
        raise HTTPException(status_code=400, detail="Invalid confirmation code")
    
    # Preserve admin
    collections = ["teams", "mentors", "evaluations", "evaluation_scores", "submissions",
                   "chat_messages", "chat_rooms", "email_logs", "sync_logs", "deadlines",
                   "rubrics", "settings", "assignment_logs", "result_release_logs", "import_sessions",
                   "round_mentor_assignments"]

    for coll in collections:
        await db[coll].drop()

    # Remove non-admin users
    await db.users.delete_many({"role": {"$ne": "admin"}})

    return {"status": "success", "message": "All data truncated except admin account"}

# ─── Export Routes (Admin) ─────────────────────────────────────────────
@app.get("/api/export/teams/excel")
async def export_teams_excel(user=Depends(require_role("admin"))):
    """Export all teams with credentials, mapping, and submissions to Excel."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    teams = await db.teams.find().to_list(1000)
    users = await db.users.find({"role": "team"}).to_list(1000)
    user_map = {u["email"]: u for u in users}
    
    # Batch fetch evaluations and submissions
    all_team_ids = [t["team_id"] for t in teams]
    eval_cursor = db.evaluations.find({"team_id": {"$in": all_team_ids}})
    evals = {}
    async for ev in eval_cursor:
        key = f"{ev['team_id']}_{ev['round_name']}"
        evals[key] = ev
    
    sub_cursor = db.submissions.find({"team_id": {"$in": all_team_ids}})
    subs = {}
    async for s in sub_cursor:
        key = f"{s['team_id']}_{s['round_name']}"
        subs[key] = s
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"
    
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = [
        "Team ID", "Team Name", "College", "Project Title", "Domain",
        "Lead Name", "Lead Email", "Lead Mobile", "Login Username", "Login Password",
        "Assigned Mentor", "Status",
        "R1 PPT Link", "R1 Submitted", "R1 Eval Status", "R1 Score",
        "R2 PPT Link", "R2 GitHub Link", "R2 Submitted", "R2 Eval Status", "R2 Score",
        "R3 PPT Link", "R3 GitHub Link", "R3 Video Link", "R3 Submitted", "R3 Eval Status", "R3 Score",
        "Members"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, team in enumerate(teams, 2):
        u = user_map.get(team.get("team_lead_email", ""))
        members_str = ", ".join([f"{m.get('name', '')} ({m.get('email', '')})" for m in team.get("members", [])])
        
        row_data = [
            team.get("team_id", ""),
            team.get("team_name", ""),
            team.get("college_name", ""),
            team.get("project_title", ""),
            team.get("project_domain", ""),
            team.get("team_lead_name", ""),
            team.get("team_lead_email", ""),
            team.get("team_lead_mobile", ""),
            u.get("username", "") if u else "",
            u.get("plain_password_temp", "N/A") if u else "",
            team.get("assigned_mentor_email", ""),
            team.get("status", ""),
        ]
        
        # Round 1
        r1_sub = subs.get(f"{team['team_id']}_Round 1", {})
        r1_eval = evals.get(f"{team['team_id']}_Round 1", {})
        row_data.extend([
            r1_sub.get("ppt_link", r1_sub.get("submission_link", "")),
            "Yes" if r1_sub else "No",
            r1_eval.get("status", "pending"),
            r1_eval.get("total_score", "") if r1_eval else "",
        ])
        
        # Round 2
        r2_sub = subs.get(f"{team['team_id']}_Round 2", {})
        r2_eval = evals.get(f"{team['team_id']}_Round 2", {})
        row_data.extend([
            r2_sub.get("ppt_link", ""),
            r2_sub.get("github_link", ""),
            "Yes" if r2_sub else "No",
            r2_eval.get("status", "pending"),
            r2_eval.get("total_score", "") if r2_eval else "",
        ])
        
        # Round 3
        r3_sub = subs.get(f"{team['team_id']}_Round 3", {})
        r3_eval = evals.get(f"{team['team_id']}_Round 3", {})
        row_data.extend([
            r3_sub.get("ppt_link", ""),
            r3_sub.get("github_link", ""),
            r3_sub.get("video_link", ""),
            "Yes" if r3_sub else "No",
            r3_eval.get("status", "pending"),
            r3_eval.get("total_score", "") if r3_eval else "",
        ])
        
        row_data.append(members_str)
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val else "")
            cell.border = thin_border
    
    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=CodHER_Teams_Export.xlsx"}
    )

@app.get("/api/export/mentors/excel")
async def export_mentors_excel(user=Depends(require_role("admin"))):
    """Export all mentors with credentials, assignments, and eval status to Excel."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    mentors = await db.mentors.find().to_list(100)
    users = await db.users.find({"role": "mentor"}).to_list(100)
    user_map = {u["email"]: u for u in users}
    
    # Batch fetch assignment data
    mentor_emails = [m["mentor_email"] for m in mentors]
    all_teams = await db.teams.find({"assigned_mentor_email": {"$in": mentor_emails}}).to_list(1000)
    
    mentor_teams_map = {}
    for t in all_teams:
        email = t.get("assigned_mentor_email", "")
        if email not in mentor_teams_map:
            mentor_teams_map[email] = []
        mentor_teams_map[email].append(t)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Mentors"
    
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = [
        "Mentor ID", "Name", "Email", "Mobile", "Login Username", "Login Password",
        "Expertise", "Organization", "Max Capacity", "Assigned Teams",
        "Assigned Team IDs", "Assigned Team Names", "Status"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, mentor in enumerate(mentors, 2):
        u = user_map.get(mentor.get("mentor_email", ""))
        assigned_teams = mentor_teams_map.get(mentor["mentor_email"], [])
        
        row_data = [
            mentor.get("mentor_id", ""),
            mentor.get("mentor_name", ""),
            mentor.get("mentor_email", ""),
            mentor.get("mentor_mobile", ""),
            u.get("username", "") if u else "",
            u.get("plain_password_temp", "N/A") if u else "",
            mentor.get("expertise", ""),
            mentor.get("organization", ""),
            mentor.get("max_team_capacity", 3),
            len(assigned_teams),
            ", ".join([t["team_id"] for t in assigned_teams]),
            ", ".join([t["team_name"] for t in assigned_teams]),
            mentor.get("status", ""),
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val else "")
            cell.border = thin_border
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=CodHER_Mentors_Export.xlsx"}
    )

@app.get("/api/export/teams/pdf")
async def export_teams_pdf(user=Depends(require_role("admin"))):
    """Export teams report as PDF."""
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    teams = await db.teams.find().to_list(1000)
    users = await db.users.find({"role": "team"}).to_list(1000)
    user_map = {u["email"]: u for u in users}
    
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, textColor=colors.HexColor("#7C3AED"))
    
    elements = []
    elements.append(Paragraph("CodHER - Teams Report", title_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Total Teams: {len(teams)}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    table_data = [["ID", "Team Name", "Lead", "Email", "Username", "Password", "Mentor", "Status"]]
    
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=9)
    
    for team in teams:
        u = user_map.get(team.get("team_lead_email", ""))
        table_data.append([
            Paragraph(str(team.get("team_id", "")), cell_style),
            Paragraph(str(team.get("team_name", "")), cell_style),
            Paragraph(str(team.get("team_lead_name", "")), cell_style),
            Paragraph(str(team.get("team_lead_email", "")), cell_style),
            Paragraph(str(u.get("username", "") if u else ""), cell_style),
            Paragraph(str(u.get("plain_password_temp", "N/A") if u else ""), cell_style),
            Paragraph(str(team.get("assigned_mentor_email", "")), cell_style),
            Paragraph(str(team.get("status", "")), cell_style),
        ])
    
    col_widths = [50, 80, 70, 120, 60, 60, 120, 50]
    t = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#7C3AED")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F6FF")]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    
    doc.build(elements)
    buf.seek(0)
    
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=CodHER_Teams_Report.pdf"}
    )

@app.get("/api/export/mentors/pdf")
async def export_mentors_pdf(user=Depends(require_role("admin"))):
    """Export mentors report as PDF."""
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    mentors = await db.mentors.find().to_list(100)
    users = await db.users.find({"role": "mentor"}).to_list(100)
    user_map = {u["email"]: u for u in users}
    
    all_teams = await db.teams.find({"assigned_mentor_email": {"$in": [m["mentor_email"] for m in mentors]}}).to_list(1000)
    mentor_teams = {}
    for t in all_teams:
        e = t.get("assigned_mentor_email", "")
        if e not in mentor_teams:
            mentor_teams[e] = []
        mentor_teams[e].append(t["team_name"])
    
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, textColor=colors.HexColor("#7C3AED"))
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=9)
    
    elements = []
    elements.append(Paragraph("CodHER - Mentors Report", title_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Total Mentors: {len(mentors)}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    table_data = [["ID", "Name", "Email", "Username", "Password", "Expertise", "Org", "Capacity", "Assigned Teams", "Status"]]
    
    for mentor in mentors:
        u = user_map.get(mentor.get("mentor_email", ""))
        assigned = mentor_teams.get(mentor["mentor_email"], [])
        table_data.append([
            Paragraph(str(mentor.get("mentor_id", "")), cell_style),
            Paragraph(str(mentor.get("mentor_name", "")), cell_style),
            Paragraph(str(mentor.get("mentor_email", "")), cell_style),
            Paragraph(str(u.get("username", "") if u else ""), cell_style),
            Paragraph(str(u.get("plain_password_temp", "N/A") if u else ""), cell_style),
            Paragraph(str(mentor.get("expertise", "")), cell_style),
            Paragraph(str(mentor.get("organization", "")), cell_style),
            Paragraph(f"{len(assigned)}/{mentor.get('max_team_capacity', 3)}", cell_style),
            Paragraph(", ".join(assigned[:5]) + ("..." if len(assigned) > 5 else ""), cell_style),
            Paragraph(str(mentor.get("status", "")), cell_style),
        ])
    
    col_widths = [45, 65, 100, 55, 55, 65, 65, 40, 100, 40]
    t = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#7C3AED")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F6FF")]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    
    doc.build(elements)
    buf.seek(0)
    
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=CodHER_Mentors_Report.pdf"}
    )

# ─── Health Check ──────────────────────────────────────────────────────
@app.get("/api/health")
async def health():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}
